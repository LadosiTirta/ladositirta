# utils/export_excel_rumus.py
# Excel dengan RUMUS — untuk keperluan internal engineer
# Sel kuning = input user, sel putih = hasil rumus (terkunci)
# Proteksi sheet dengan password agar rumus tidak mudah diubah
#
# Sheet: 1_Input | 2_Tekan_Tarik | 3_Lateral_Broms | 4_Ringkasan

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PASSWORD = "admin123"   # ganti sesuai kebutuhan

# ── Style dasar ──────────────────────────────────────────────
F_HDR  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
F_SUB  = Font(name="Arial", bold=True, color="1A5376", size=10)
F_BODY = Font(name="Arial", size=10)
F_LOCK = Font(name="Arial", size=10, color="1A1A1A")
F_NOTE = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HDR   = PatternFill("solid", fgColor="1A5376")
FILL_INPUT = PatternFill("solid", fgColor="FFF9C4")   # kuning muda = input
FILL_HASIL = PatternFill("solid", fgColor="EBF5FB")   # biru muda = rumus
FILL_SUB   = PatternFill("solid", fgColor="D6EAF8")
FILL_OK    = PatternFill("solid", fgColor="D5F5E3")   # hijau = ringkasan
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left",   vertical="center")
AL_R = Alignment(horizontal="right",  vertical="center")

_s = Side(style="thin", color="BBBBBB")
BRD = Border(left=_s, right=_s, top=_s, bottom=_s)

UNLOCKED = Protection(locked=False)
LOCKED   = Protection(locked=True)

FMT2 = '#,##0.00'
FMT3 = '#,##0.000'
FMT0 = '#,##0'


def _h(ws, r, c, val, fill=None, font=None, align=None, fmt=None,
        merge=None, prot=None):
    """Set sel dengan cepat."""
    cell = ws.cell(r, c, val)
    if fill:  cell.fill   = fill
    if font:  cell.font   = font
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt
    if merge: ws.merge_cells(start_row=r, start_column=c,
                              end_row=r, end_column=merge)
    if prot:  cell.protection = prot
    cell.border = BRD
    return cell


def _header(ws, r, cols):
    for i, txt in enumerate(cols):
        _h(ws, r, i+1, txt, FILL_HDR, F_HDR, AL_C)
    ws.row_dimensions[r].height = 28


def _col_w(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════
# SHEET 1 — INPUT
# ════════════════════════════════════════════════════════════
def _sheet_input(ws, param_tiang, df_tanah):
    ws.title = "1_Input"
    _col_w(ws, {"A":5,"B":22,"C":13,"D":13,"E":11,
                "F":13,"G":11,"H":11,"I":12})

    # ── Judul ──
    _h(ws,1,1,"INPUT PERHITUNGAN KAPASITAS PONDASI TIANG",
       FILL_HDR, F_HDR, AL_C, merge=9)
    ws.row_dimensions[1].height = 24

    _h(ws,2,1,"Sel KUNING = input yang bisa diubah  |  "
              "Sel BIRU = hasil rumus (terkunci)",
       None, F_NOTE, AL_L, merge=9)

    # ── Blok parameter tiang (baris 4–14) ──
    _h(ws,4,1,"PARAMETER TIANG", FILL_HDR, F_HDR, AL_C, merge=4)
    _h(ws,4,5,"NILAI", FILL_HDR, F_HDR, AL_C)
    _h(ws,4,6,"SATUAN", FILL_HDR, F_HDR, AL_C)

    params = [
        ("Diameter / lebar tiang (D)",  param_tiang["diameter"],         "m"),
        ("Kedalaman tiang (L)",          param_tiang["kedalaman"],         "m"),
        ("Muka air tanah (MAT)",          param_tiang["muka_air"],          "m"),
        ("Safety factor tekan (SF_c)",   param_tiang["sf_tekan"],          "—"),
        ("Safety factor tarik (SF_t)",   param_tiang["sf_tarik"],          "—"),
        ("fc' beton (MPa)",              param_tiang.get("fc_prime", 33.2),"MPa"),
    ]
    # Tipe tiang → dropdown
    tipe_opsi = ('"Displacement (pancang),Non-displacement (boredpile)"')
    dv_tipe = DataValidation(type="list", formula1=tipe_opsi, allow_blank=False)
    ws.add_data_validation(dv_tipe)

    _h(ws,5,1,"Tipe tiang", FILL_SUB, F_SUB, AL_L, merge=4)
    tipe_val = "Displacement (pancang)" if param_tiang["is_displacement"] else "Non-displacement (boredpile)"
    c_tipe = _h(ws,5,5, tipe_val, FILL_INPUT, F_BODY, AL_L, prot=UNLOCKED)
    dv_tipe.add(c_tipe)
    _h(ws,5,6,"—", FILL_WHITE, F_BODY, AL_C)

    for i,(lbl,val,sat) in enumerate(params):
        r = 6 + i
        _h(ws, r, 1, lbl,  FILL_SUB,   F_SUB,  AL_L, merge=4)
        _h(ws, r, 5, val,  FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        _h(ws, r, 6, sat,  FILL_WHITE, F_BODY, AL_C)

    # Sel referensi cepat (label di kolom 8-9)
    refs = [
        ("D (m)",   "=E6"),  ("L (m)",   "=E7"),
        ("MAT (m)", "=E8"),  ("SF_c",    "=E9"),
        ("SF_t",    "=E10"), ("fc' MPa", "=E11"),
    ]
    _h(ws,4,8,"Nama sel", FILL_HDR, F_HDR, AL_C)
    _h(ws,4,9,"Referensi", FILL_HDR, F_HDR, AL_C)
    for i,(nm,ref) in enumerate(refs):
        _h(ws,5+i,8, nm,  FILL_SUB,   F_SUB,  AL_L)
        _h(ws,5+i,9, ref, FILL_HASIL, F_LOCK, AL_C, prot=LOCKED)

    # ── Blok data tanah (mulai baris 16) ──
    _h(ws,15,1,"DATA TANAH PER LAPISAN", FILL_HDR, F_HDR, AL_C, merge=9)
    _h(ws,16,1,"Catatan: Isi Cu=0 untuk pasir, φ=0 untuk lempung",
       None, F_NOTE, AL_L, merge=9)

    _header(ws, 17, ["No.","Jenis Tanah","z Atas\n(m)","z Bawah\n(m)",
                      "SPT-N","Cu\n(kPa)","φ\n(°)","γ\n(kN/m³)","Tebal\n(m)"])

    from utils.input_handler import KOLOM_TANAH as KT
    n_lap = len(df_tanah)
    for i, (_, baris) in enumerate(df_tanah.iterrows()):
        r = 18 + i
        _h(ws,r,1, i+1,                             FILL_WHITE, F_BODY, AL_C, prot=UNLOCKED)
        _h(ws,r,2, str(baris[KT["jenis"]]),          FILL_INPUT, F_BODY, AL_L, prot=UNLOCKED)
        _h(ws,r,3, float(baris[KT["z_atas"]]),      FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        _h(ws,r,4, float(baris[KT["z_bawah"]]),     FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        _h(ws,r,5, int(baris[KT["spt"]]),            FILL_INPUT, F_BODY, AL_R, prot=UNLOCKED)
        _h(ws,r,6, float(baris[KT["cu"]]),           FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        _h(ws,r,7, float(baris[KT["phi"]]),          FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        _h(ws,r,8, float(baris[KT["gamma"]]),        FILL_INPUT, F_BODY, AL_R, fmt=FMT2, prot=UNLOCKED)
        # Tebal = rumus
        _h(ws,r,9, f"=D{r}-C{r}", FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)

    # Baris info jumlah lapisan
    r_end = 18 + n_lap
    _h(ws,r_end,1, f"Jumlah lapisan: {n_lap}", None, F_NOTE, AL_L, merge=9)

    # Simpan posisi penting sebagai named range (via komentar sel)
    ws["A1"].comment = None
    # Simpan n_lap ke sel tersembunyi untuk dipakai sheet lain
    ws.cell(r_end+2, 9).value = n_lap
    ws.cell(r_end+2, 9).font  = Font(color="FFFFFF", size=8)  # tersembunyi (putih)

    # Proteksi sheet — biarkan sel kuning bisa diedit
    ws.protection.sheet    = True
    ws.protection.password = PASSWORD
    ws.protection.enable()


# ════════════════════════════════════════════════════════════
# SHEET 2 — TEKAN & TARIK (rumus)
# ════════════════════════════════════════════════════════════
def _sheet_tekan(ws, df_tanah, param_tiang):
    ws.title = "2_Tekan_Tarik"
    n = len(df_tanah)
    _col_w(ws, {"A":5,"B":16,"C":11,"D":11,"E":11,"F":11,
                "G":10,"H":10,"I":10,"J":12})

    _h(ws,1,1,"PERHITUNGAN DAYA DUKUNG TEKAN & TARIK",
       FILL_HDR, F_HDR, AL_C, merge=10)
    _h(ws,2,1,"Semua rumus merujuk ke sheet '1_Input' — ubah input di sana",
       None, F_NOTE, AL_L, merge=10)

    # ── Tabel tegangan efektif + skin friction ──
    _h(ws,4,1,"DISTRIBUSI SKIN FRICTION PER LAPISAN",
       FILL_SUB, F_SUB, AL_C, merge=10)
    _header(ws, 5, ["No.","Jenis Tanah","z Tengah\n(m)","Tebal\n(m)",
                     "γ\n(kN/m³)","σv\n(kPa)","σ'v\n(kPa)","α / β","fs\n(kPa)","Qs\n(kN)"])

    # Referensi sheet input: data lapisan mulai baris 18
    INP = "'1_Input'"  # nama sheet input

    for i in range(n):
        r   = 6 + i
        ri  = 18 + i  # baris di sheet input

        # z tengah, tebal, gamma — langsung dari Input
        z_atas  = f"{INP}!C{ri}"
        z_bawah = f"{INP}!D{ri}"
        gamma   = f"{INP}!H{ri}"
        cu      = f"{INP}!F{ri}"
        phi     = f"{INP}!G{ri}"
        spt     = f"{INP}!E{ri}"
        tebal   = f"{INP}!I{ri}"  # rumus =D-C di sheet input

        z_tengah_f = f"=({z_atas}+{z_bawah})/2"
        # σv di tengah lapisan (akumulasi dari atas)
        if i == 0:
            sv_f  = f"={gamma}*({z_tengah_f[1:]})"          # tanpa '='
            sv_f  = f"={gamma}*({z_atas}+{tebal}/2)"
        else:
            # akumulasi σv baris sebelumnya + γ*(tebal/2)
            sv_f  = f"=J{r-1}/1+{gamma}*{tebal}"            # approx via col J
            # Cara lebih tepat: jumlah γ*tebal lapisan di atasnya
            sv_parts = "+".join(
                f"{INP}!H{18+j}*{INP}!I{18+j}" for j in range(i)
            )
            sv_f = f"=({sv_parts})+{gamma}*({tebal}/2)"

        # Tekanan air pori
        mat_ref = f"{INP}!E8"
        u_f   = f"=IF({z_tengah_f[1:]}>={mat_ref},9.81*({z_tengah_f[1:]}-{mat_ref}),0)"
        sv_eff_f = f"=MAX({sv_f[1:]}-({u_f[1:]}),0)"

        # α (lempung) — Tomlinson 2008
        # β (pasir)  — Burland 1973
        # tipe displacement
        is_disp_ref = f"={INP}!E5"
        alpha_f = (
            f'=IF({cu}>0,'
            f'IF({INP}!E5="Displacement (pancang)",'
            f'IF({cu}<=25,1.0,IF({cu}<=50,0.9,IF({cu}<=75,0.75,'
            f'IF({cu}<=100,0.55,IF({cu}<=150,0.45,0.35))))),'
            f'IF({cu}<=25,0.9,IF({cu}<=50,0.75,IF({cu}<=75,0.6,'
            f'IF({cu}<=100,0.45,IF({cu}<=150,0.35,0.25)))))),0)'
        )
        # β = K*tan(δ) — disederhanakan
        kp_f  = f'=IF({INP}!E5="Displacement (pancang)",1.0+0.5*MIN({z_tengah_f[1:]}/20,1),0.5+0.2*MIN({z_tengah_f[1:]}/20,1))'
        phi_r = f"=RADIANS({phi})"
        beta_f= f"=MIN(({kp_f[1:]})*TAN(0.75*RADIANS({phi})),0.5)"

        # Pilih metode: lempung (Cu>0) atau pasir
        ab_f  = f"=IF({cu}>0,{alpha_f[1:]},{beta_f[1:]})"

        # fs = α*Cu (lempung) atau β*σ'v (pasir)
        # Rumus σ'v lengkap untuk fs
        sv_eff_clean = sv_eff_f[1:]
        fs_f  = (f"=IF({cu}>0,"
                 f"({alpha_f[1:]})*{cu},"
                 f"MAX(({beta_f[1:]})*({sv_eff_clean}),0))")

        # Keliling tiang: π*D (bulat) atau 4*D (kotak)
        D_ref  = f"{INP}!E6"
        # Gunakan 4*D sebagai default (square pile); bisa diubah
        kel_f  = f"=4*{D_ref}"

        Qs_f   = f"=({fs_f[1:]})*({kel_f[1:]})*{tebal}"

        _h(ws,r,1, i+1,                    FILL_WHITE, F_BODY, AL_C)
        _h(ws,r,2, f"={INP}!B{ri}",        FILL_HASIL, F_LOCK, AL_L, prot=LOCKED)
        _h(ws,r,3, z_tengah_f,             FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,4, f"={tebal}",            FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,5, f"={gamma}",            FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,6, sv_f,                   FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,7, sv_eff_f,               FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,8, ab_f,                   FILL_HASIL, F_LOCK, AL_R, fmt=FMT3, prot=LOCKED)
        _h(ws,r,9, fs_f,                   FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)
        _h(ws,r,10,Qs_f,                   FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)

    # ── Total Qskin ──
    r_tot = 6 + n
    Qskin_range = f"J6:J{r_tot-1}"
    _h(ws,r_tot,1,"TOTAL ΣQskin (kN)", FILL_OK, F_SUB, AL_L, merge=9)
    _h(ws,r_tot,10, f"=SUM({Qskin_range})", FILL_OK,
       Font(name="Arial",bold=True,size=11,color="1E8449"), AL_R, fmt=FMT2, prot=LOCKED)

    # ── End bearing Qpoint ──
    r_ep = r_tot + 2
    _h(ws,r_ep,1,"END BEARING (Qpoint)", FILL_HDR, F_HDR, AL_C, merge=10)

    # SPT rata-rata 4D di ujung — pakai lapisan terakhir sebagai pendekatan
    ri_last = 18 + n - 1
    D_ref   = f"{INP}!E6"
    L_ref   = f"{INP}!E7"
    spt_last= f"{INP}!E{ri_last}"
    # Nq interpolasi dari phi lapisan ujung (pendekatan sederhana)
    phi_ujung = f"{INP}!G{ri_last}"
    nq_f    = (f"=IF({phi_ujung}<=10,3,IF({phi_ujung}<=20,10,"
               f"IF({phi_ujung}<=25,20,IF({phi_ujung}<=30,40,"
               f"IF({phi_ujung}<=35,80,IF({phi_ujung}<=40,150,250))))))")
    Ab_f    = f"=PI()/4*{D_ref}^2"  # tiang bulat; untuk kotak gunakan =D^2
    # σ'v ujung tiang (akumulasi semua lapisan)
    sv_all  = "+".join(f"{INP}!H{18+j}*{INP}!I{18+j}" for j in range(n))
    sv_ujung= f"=MAX({sv_all}-9.81*MAX({L_ref}-{INP}!E8,0),0)"
    # qp SPT (Meyerhof): 0.4*N*(L/D) maks 4*N
    qp_spt_f= (f"=MIN(0.4*{spt_last}*({L_ref}/{D_ref}),"
               f"4*{spt_last})")
    Qpoint_nq_f = f"=({nq_f[1:]})*({sv_ujung[1:]})*({Ab_f[1:]})"
    Qpoint_spt_f= f"=({qp_spt_f[1:]})*({Ab_f[1:]})"
    Qpoint_f    = f"=MIN({Qpoint_nq_f[1:]},{Qpoint_spt_f[1:]})"

    info_ep = [
        ("Luas ujung Ab (m²)",      Ab_f),
        ("Nq (Meyerhof)",           nq_f),
        ("σ'v ujung (kPa)",         sv_ujung),
        ("Qpoint — cara Nq (kN)",   Qpoint_nq_f),
        ("Qpoint — cara SPT (kN)",  Qpoint_spt_f),
        ("Qpoint dipakai (kN)",     Qpoint_f),
    ]
    for i,(lbl,frm) in enumerate(info_ep):
        _h(ws,r_ep+1+i,1, lbl, FILL_SUB, F_SUB, AL_L, merge=9)
        _h(ws,r_ep+1+i,10, frm, FILL_HASIL, F_LOCK, AL_R, fmt=FMT2, prot=LOCKED)

    # ── Kapasitas total ──
    r_kap = r_ep + len(info_ep) + 2
    _h(ws,r_kap,1,"KAPASITAS TOTAL", FILL_HDR, F_HDR, AL_C, merge=10)
    r_kap += 1

    Qskin_sel  = f"J{r_tot}"
    Qpoint_sel = f"J{r_ep+len(info_ep)}"
    SF_c_ref   = f"{INP}!E9"
    SF_t_ref   = f"{INP}!E10"
    fc_ref     = f"{INP}!E11"

    fr_tarik  = "0.85"  # displacement; akan 0.70 untuk boredpile (disederhanakan)
    Qult_c_f  = f"={Qpoint_sel}+{Qskin_sel}"
    Qult_t_f  = f"={Qskin_sel}*{fr_tarik}"
    Qijin_c_f = f"={Qult_c_f[1:]}/{SF_c_ref}"
    Qijin_t_f = f"={Qult_t_f[1:]}/{SF_t_ref}"
    Pn_f      = f"=0.65*0.85*{fc_ref}*1000*(PI()/4*{D_ref}^2)*0.85"

    kapasitas = [
        ("Qultimate tekan (kN)",    Qult_c_f),
        ("Qijin tekan (kN)",        Qijin_c_f),
        ("Qultimate tarik (kN)",    Qult_t_f),
        ("Qijin tarik (kN)",        Qijin_t_f),
        ("Kapasitas struktur Pn (kN)", Pn_f),
    ]
    for i,(lbl,frm) in enumerate(kapasitas):
        bold = "ijin" in lbl.lower()
        fnt  = Font(name="Arial",bold=bold,size=10,
                    color="1E8449" if bold else "1A1A1A")
        _h(ws,r_kap+i,1, lbl, FILL_OK if bold else FILL_HASIL,
           fnt, AL_L, merge=9)
        _h(ws,r_kap+i,10, frm,
           FILL_OK if bold else FILL_HASIL,
           fnt, AL_R, fmt=FMT2, prot=LOCKED)

    ws.protection.sheet    = True
    ws.protection.password = PASSWORD
    ws.protection.enable()


# ════════════════════════════════════════════════════════════
# SHEET 3 — GAYA LATERAL BROMS (rumus)
# ════════════════════════════════════════════════════════════
def _sheet_lateral(ws):
    ws.title = "3_Lateral_Broms"
    _col_w(ws, {"A":30,"B":20,"C":12,"D":20})

    _h(ws,1,1,"GAYA LATERAL — BROMS (1964)",
       FILL_HDR, F_HDR, AL_C, merge=4)
    _h(ws,2,1,"Parameter tiang D, L diambil dari '1_Input'",
       None, F_NOTE, AL_L, merge=4)

    INP   = "'1_Input'"
    D_ref = f"{INP}!E6"
    L_ref = f"{INP}!E7"

    # Input gaya lateral
    _h(ws,4,1,"INPUT GAYA LATERAL", FILL_HDR, F_HDR, AL_C, merge=4)
    inp_lat = [
        ("Gaya lateral H (kN)",     80.0),
        ("Eksentrisitas e (m)",      0.5),
        ("Cu rata-rata (kPa)",       30.0),
        ("φ rata-rata (°)",          0.0),
        ("γ tanah rata-rata (kN/m³)",17.0),
        ("MAT (m) — dari Input",    f"={INP}!E8"),
    ]
    for i,(lbl,val) in enumerate(inp_lat):
        _h(ws,5+i,1, lbl, FILL_SUB, F_SUB, AL_L, merge=2)
        is_frm = isinstance(val, str) and val.startswith("=")
        _h(ws,5+i,3, val,
           FILL_HASIL if is_frm else FILL_INPUT,
           F_LOCK if is_frm else F_BODY,
           AL_R, fmt=FMT2,
           prot=LOCKED if is_frm else UNLOCKED)

    # Referensi sel input lateral
    H_r  = "C5"   # gaya lateral
    e_r  = "C6"   # eksentrisitas
    Cu_r = "C7"   # Cu rata-rata
    phi_r= "C8"   # phi rata-rata
    gam_r= "C9"   # gamma

    # Dropdown kondisi kepala
    dv = DataValidation(type="list",
                        formula1='"Kepala Bebas,Kepala Jepit"',
                        allow_blank=False)
    ws.add_data_validation(dv)
    _h(ws,11,1,"Kondisi kepala tiang", FILL_SUB, F_SUB, AL_L, merge=2)
    c_kep = _h(ws,11,3,"Kepala Bebas", FILL_INPUT, F_BODY, AL_C, prot=UNLOCKED)
    dv.add(c_kep)
    kondisi_ref = "C11"

    # Hitung hasil (rumus)
    _h(ws,13,1,"HASIL PERHITUNGAN BROMS",
       FILL_HDR, F_HDR, AL_C, merge=4)

    # p_ult lempung: 9*Cu*D
    pult_lempung = f"=9*{Cu_r}*{D_ref}"
    # Hu lempung kepala bebas: iterasi → pendekatan Fp*(L/3)/(L+e)
    # Fp = 0.5*p_ult*L² ... pendekatan
    Hu_lemb_f = (f"=IF({kondisi_ref}=\"Kepala Bebas\","
                 f"({pult_lempung[1:]})*({L_ref}^2/2)/({L_ref}+{e_r}),"
                 f"({pult_lempung[1:]})*({L_ref}^2/2)/({L_ref}+{e_r})*1.5)")

    # Kp pasir
    Kp_f   = f"=(1+SIN(RADIANS({phi_r})))/(1-SIN(RADIANS({phi_r})))"
    gam_eff= f"=MAX({gam_r}-9.81,8)"
    Fp_f   = f"=0.5*3*({Kp_f[1:]})*({gam_eff[1:]})*{L_ref}^2*{D_ref}"
    Hu_pasir_f = (f"=IF({kondisi_ref}=\"Kepala Bebas\","
                  f"({Fp_f[1:]})*({L_ref}/3)/({L_ref}+{e_r}),"
                  f"({Fp_f[1:]})*({L_ref}/3)/({L_ref}+{e_r})*1.5)")

    # Pilih lempung atau pasir berdasarkan Cu
    Hu_f   = f"=IF({Cu_r}>0,{Hu_lemb_f[1:]},{Hu_pasir_f[1:]})"
    SF_lat = 2.5
    Hijin_f= f"={Hu_f[1:]}/{SF_lat}"
    Mmax_f = f"={Hu_f[1:]}*({e_r}+1.5*{D_ref}+{Hu_f[1:]}/(9*MAX({Cu_r},1)*{D_ref}))"

    # EI tiang (beton): Ec=4700*sqrt(fc')*1000 kN/m² × Ip
    fc_ref  = f"='{INP[1:-1]}'!E11"
    EI_f    = f"=4700*SQRT({INP}!E11)*1000*(PI()/64*{D_ref}^4)"
    kh_f    = f"=IF({Cu_r}>0,67*{Cu_r},2000*{Kp_f[1:]})"
    beta_f  = f"=({kh_f[1:]}*{D_ref}/(4*{EI_f[1:]}))^0.25"
    LB_f    = f"={beta_f[1:]}*{L_ref}"
    defl_f  = (f"=IF({Cu_r}>0,"
               f"IF({LB_f[1:]}>=2,{H_r}*{beta_f[1:]}/({kh_f[1:]}*{D_ref})*(2+{beta_f[1:]}*{e_r}),{H_r}*({L_ref}+{e_r})^3/(3*{EI_f[1:]})),"
               f"IF({LB_f[1:]}>=5,2.435*{H_r}*({EI_f[1:]}/{kh_f[1:]})^0.6/{EI_f[1:]},{H_r}*({L_ref}+{e_r})^3/(3*{EI_f[1:]}))"
               f")*1000")  # mm
    kontrol_f = f'=IF({H_r}<={Hijin_f[1:]},"OK ✓","TIDAK OK ✗")'

    hasil_lat = [
        ("EI tiang (kN·m²)",           EI_f),
        ("p_ult lempung (kN/m)",        pult_lempung),
        ("Kp pasir",                    Kp_f),
        ("Hu ultimit (kN)",             Hu_f),
        ("Hijin = Hu / 2.5 (kN)",      Hijin_f),
        ("Mmax (kN·m)",                 Mmax_f),
        ("Defleksi kepala y₀ (mm)",     defl_f),
        ("Kontrol H ≤ Hijin",           kontrol_f),
    ]
    for i,(lbl,frm) in enumerate(hasil_lat):
        bold = "Hijin" in lbl or "Kontrol" in lbl
        fnt  = Font(name="Arial", bold=bold, size=10,
                    color="1E8449" if bold else "1A1A1A")
        _h(ws,14+i,1, lbl, FILL_OK if bold else FILL_HASIL,
           fnt, AL_L, merge=3)
        _h(ws,14+i,4, frm,
           FILL_OK if bold else FILL_HASIL,
           fnt, AL_R, fmt=FMT2 if "Kontrol" not in lbl else "@",
           prot=LOCKED)

    ws.protection.sheet    = True
    ws.protection.password = PASSWORD
    ws.protection.enable()


# ════════════════════════════════════════════════════════════
# SHEET 4 — RINGKASAN
# ════════════════════════════════════════════════════════════
def _sheet_ringkasan(ws, n_lap):
    ws.title = "4_Ringkasan"
    _col_w(ws, {"A":30,"B":20,"C":12,"D":20})

    _h(ws,1,1,"RINGKASAN HASIL PERHITUNGAN",
       FILL_HDR, F_HDR, AL_C, merge=4)
    _h(ws,2,1,"Semua nilai merujuk langsung dari sheet perhitungan",
       None, F_NOTE, AL_L, merge=4)

    S2 = "'2_Tekan_Tarik'"
    S3 = "'3_Lateral_Broms'"
    S1 = "'1_Input'"

    # Hitung baris hasil di sheet 2
    # r_tot = 6+n_lap (baris total Qskin)
    # r_ep  = r_tot+2
    # info_ep = 6 baris → Qpoint di r_ep+6
    # r_kap = r_ep+8 → kapasitas mulai r_kap
    r_tot = 6 + n_lap
    r_ep  = r_tot + 2
    r_kap = r_ep + 6 + 2  # +len(info_ep)+2

    rows = [
        ("DAYA DUKUNG AKSIAL",        None),
        ("Qpoint — End Bearing (kN)", f"={S2}!J{r_ep+6}"),
        ("ΣQskin — Skin Friction (kN)",f"={S2}!J{r_tot}"),
        ("Qult tekan (kN)",            f"={S2}!J{r_kap}"),
        ("Qijin tekan (kN)",           f"={S2}!J{r_kap+1}"),
        ("Qult tarik (kN)",            f"={S2}!J{r_kap+2}"),
        ("Qijin tarik (kN)",           f"={S2}!J{r_kap+3}"),
        ("Kapasitas struktur Pn (kN)", f"={S2}!J{r_kap+4}"),
        ("GAYA LATERAL",               None),
        ("Hu ultimit (kN)",            f"={S3}!D17"),
        ("Hijin lateral (kN)",         f"={S3}!D18"),
        ("Momen maks Mmax (kN·m)",     f"={S3}!D19"),
        ("Defleksi kepala y₀ (mm)",    f"={S3}!D20"),
        ("Kontrol H ≤ Hijin",          f"={S3}!D21"),
        ("PARAMETER INPUT",            None),
        ("Diameter tiang D (m)",       f"={S1}!E6"),
        ("Kedalaman tiang L (m)",      f"={S1}!E7"),
        ("Muka air tanah MAT (m)",     f"={S1}!E8"),
        ("Safety factor tekan",        f"={S1}!E9"),
        ("Safety factor tarik",        f"={S1}!E10"),
    ]

    r = 4
    for lbl, frm in rows:
        if frm is None:
            # Sub-heading
            _h(ws,r,1, lbl, FILL_HDR, F_HDR, AL_C, merge=4)
        else:
            is_key = any(k in lbl for k in ["ijin","Hijin","Kontrol"])
            fnt = Font(name="Arial", bold=is_key, size=10,
                       color="1E8449" if is_key else "1A1A1A")
            fill = FILL_OK if is_key else FILL_HASIL
            _h(ws,r,1, lbl, FILL_SUB, F_SUB, AL_L, merge=3)
            _h(ws,r,4, frm, fill, fnt, AL_R,
               fmt=FMT2 if "Kontrol" not in lbl else "@",
               prot=LOCKED)
        r += 1

    ws.protection.sheet    = True
    ws.protection.password = PASSWORD
    ws.protection.enable()


# ════════════════════════════════════════════════════════════
# FUNGSI UTAMA
# ════════════════════════════════════════════════════════════
def buat_excel_rumus(
    param_tiang: dict,
    df_tanah,
    nama_proyek: str = "Proyek",
    nama_konsultan: str = "",
) -> io.BytesIO:
    """
    Membuat Excel dengan RUMUS untuk keperluan internal engineer.
    Sheet terkunci dengan password (default: 'admin123').
    Sel kuning = input yang bisa diubah.
    Sel biru/hijau = hasil rumus (terkunci).

    Mengembalikan BytesIO.
    """
    wb = Workbook()
    wb.remove(wb.active)

    n = len(df_tanah)
    ws1 = wb.create_sheet(); _sheet_input(ws1, param_tiang, df_tanah)
    ws2 = wb.create_sheet(); _sheet_tekan(ws2, df_tanah, param_tiang)
    ws3 = wb.create_sheet(); _sheet_lateral(ws3)
    ws4 = wb.create_sheet(); _sheet_ringkasan(ws4, n)

    wb.active = ws1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
