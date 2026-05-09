# utils/input_handler.py
# Penanganan input data tanah dan parameter tiang
# Mendukung dua cara input: manual (st.data_editor) dan upload Excel

import streamlit as st
import pandas as pd
import numpy as np

# =============================================================
# KONSTANTA NAMA KOLOM TABEL DATA TANAH
# =============================================================
KOLOM_TANAH = {
    "no":       "No.",
    "z_atas":   "Kedalaman Atas (m)",
    "z_bawah":  "Kedalaman Bawah (m)",
    "jenis":    "Jenis Tanah",
    "spt":      "SPT-N (pukulan)",
    "cu":       "Cu (kPa)",
    "phi":      "φ (°)",
    "gamma":    "γ (kN/m³)",
}

JENIS_TANAH_OPSI = [
    "Lempung sangat lunak",
    "Lempung lunak",
    "Lempung sedang",
    "Lempung kaku",
    "Lempung sangat kaku",
    "Pasir lepas",
    "Pasir sedang",
    "Pasir padat",
    "Lanau",
    "Kerikil",
]

# Data contoh default untuk memudahkan user memulai
DATA_TANAH_DEFAULT = pd.DataFrame({
    KOLOM_TANAH["no"]:     [1, 2, 3, 4],
    KOLOM_TANAH["z_atas"]: [0.0, 3.0, 8.0, 15.0],
    KOLOM_TANAH["z_bawah"]:[3.0, 8.0, 15.0, 25.0],
    KOLOM_TANAH["jenis"]:  ["Lempung lunak", "Lempung sedang", "Pasir sedang", "Pasir padat"],
    KOLOM_TANAH["spt"]:    [4, 10, 22, 40],
    KOLOM_TANAH["cu"]:     [20.0, 40.0, 0.0, 0.0],
    KOLOM_TANAH["phi"]:    [0.0, 0.0, 30.0, 35.0],
    KOLOM_TANAH["gamma"]:  [16.0, 17.0, 18.0, 19.0],
})


def _buat_template_excel_lengkap() -> bytes:
    """
    Membuat template Excel form input data tanah yang lengkap:
    - Baris header berwarna
    - Contoh data 4 lapisan
    - Baris & kolom batas dengan keterangan
    - Petunjuk pengisian
    Mengembalikan bytes siap download.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Data_Tanah"

    # -- Style --
    BIRU   = "1A5376"; KUNING = "FFF9C4"; HIJAU = "E8F8E8"
    ABU    = "F2F3F4"; MERAH  = "FADBD8"; PUTIH = "FFFFFF"
    f_hdr  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    f_body = Font(name="Arial", size=10)
    f_note = Font(name="Arial", size=9, italic=True, color="555555")
    f_bts  = Font(name="Arial", bold=True, size=9, color="922B21")
    al_c   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_l   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    s      = Side(style="thin", color="AAAAAA")
    brd    = Border(left=s, right=s, top=s, bottom=s)

    def sel(r, c, val=None, bg=None, font=None, align=None, w=None):
        cell = ws.cell(r, c, val)
        if bg:    cell.fill   = PatternFill("solid", fgColor=bg)
        if font:  cell.font   = font
        if align: cell.alignment = align
        cell.border = brd
        return cell

    # -- Baris 1: judul --
    ws.merge_cells("A1:I1")
    t = ws["A1"]; t.value = "FORM INPUT DATA TANAH — Program Kapasitas Pondasi Tiang"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=BIRU)
    t.alignment = al_c
    ws.row_dimensions[1].height = 24

    # -- Baris 2: petunjuk --
    ws.merge_cells("A2:I2")
    p = ws["A2"]
    p.value = ("Isi data tanah mulai baris 5. Sel KUNING = wajib diisi. "
               "Cu=0 untuk pasir, φ=0 untuk lempung. "
               "JANGAN hapus/geser kolom. Simpan sebagai .xlsx lalu upload.")
    p.font = f_note; p.alignment = al_l
    p.fill = PatternFill("solid", fgColor="FEF9E7")
    ws.row_dimensions[2].height = 18

    # -- Baris 3: header kolom --
    headers = [
        ("No.", 5), ("Jenis Tanah", 20),
        ("Kedalaman Atas (m)", 14), ("Kedalaman Bawah (m)", 14),
        ("SPT-N (pukulan)", 12), ("Cu (kPa)", 10),
        ("φ (°)", 8), ("γ (kN/m³)", 10),
        ("Catatan", 20),
    ]
    for i, (nama, lebar) in enumerate(headers):
        c = i + 1
        cell = ws.cell(3, c, nama)
        cell.font  = f_hdr
        cell.fill  = PatternFill("solid", fgColor=BIRU)
        cell.alignment = al_c
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = lebar
    ws.row_dimensions[3].height = 30

    # -- Baris 4: sub-header satuan --
    satuan = ["—", "pilih jenis tanah", "m", "m", "pukulan", "kPa", "derajat", "kN/m³", "opsional"]
    for i, sat in enumerate(satuan):
        cell = ws.cell(4, i+1, sat)
        cell.font = Font(name="Arial", size=9, italic=True, color="555555")
        cell.fill = PatternFill("solid", fgColor="D6EAF8")
        cell.alignment = al_c; cell.border = brd
    ws.row_dimensions[4].height = 16

    # -- Data contoh (baris 5-8) --
    contoh = [
        (1, "Lempung lunak",  0.0,  3.0,  4, 20.0,  0.0, 16.0, "Lapisan permukaan"),
        (2, "Lempung sedang", 3.0,  8.0, 10, 40.0,  0.0, 17.0, ""),
        (3, "Pasir sedang",   8.0, 15.0, 22,  0.0, 30.0, 18.0, "Cu=0 karena pasir"),
        (4, "Pasir padat",   15.0, 25.0, 40,  0.0, 35.0, 19.0, ""),
    ]
    for i, baris in enumerate(contoh):
        r = 5 + i
        for j, val in enumerate(baris):
            bg = KUNING if j in [1,2,3,4,5,6,7] else ABU
            cell = ws.cell(r, j+1, val)
            cell.font = f_body; cell.alignment = al_c if j != 1 else al_l
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = brd
        ws.row_dimensions[r].height = 18

    # -- Baris batas bawah (baris 9) --
    baris_batas = 9
    ws.merge_cells(f"A{baris_batas}:I{baris_batas}")
    bb = ws[f"A{baris_batas}"]
    bb.value = ("⬆ BATAS BAWAH DATA — Jika ingin menambah lapisan, "
                "INSERT ROW di atas baris ini (klik no. baris → Insert). "
                "Program membaca otomatis hingga baris terakhir berisi data.")
    bb.font  = f_bts
    bb.fill  = PatternFill("solid", fgColor=MERAH)
    bb.alignment = al_l; bb.border = brd
    ws.row_dimensions[baris_batas].height = 20

    # -- Kolom batas kanan (kolom J) --
    for r in range(3, baris_batas + 1):
        cell = ws.cell(r, 10,
            "◄ BATAS KANAN — jangan tambah kolom di kiri ini" if r == 3 else "")
        cell.font = f_bts if r == 3 else f_note
        cell.fill = PatternFill("solid", fgColor=MERAH)
        cell.alignment = al_c; cell.border = brd
    ws.column_dimensions["J"].width = 32

    # -- Sheet 2: Petunjuk --
    ws2 = wb.create_sheet("Petunjuk")
    petunjuk = [
        ("PETUNJUK PENGISIAN FORM DATA TANAH", True, BIRU, "FFFFFF"),
        ("", False, PUTIH, "000000"),
        ("1. JENIS TANAH", True, "D6EAF8", BIRU),
        ("   Pilih dari: Lempung sangat lunak / lunak / sedang / kaku / sangat kaku", False, PUTIH, "000000"),
        ("   atau: Pasir lepas / sedang / padat, Lanau, Kerikil", False, PUTIH, "000000"),
        ("", False, PUTIH, "000000"),
        ("2. KEDALAMAN", True, "D6EAF8", BIRU),
        ("   z atas lap. 1 = 0.0 (permukaan tanah)", False, PUTIH, "000000"),
        ("   z bawah lap. 1 = z atas lap. 2 (harus sambung, tidak boleh ada gap)", False, PUTIH, "000000"),
        ("", False, PUTIH, "000000"),
        ("3. SPT-N", True, "D6EAF8", BIRU),
        ("   Nilai pukulan per 30cm. Isi 0 jika tidak ada data SPT.", False, PUTIH, "000000"),
        ("", False, PUTIH, "000000"),
        ("4. Cu dan φ", True, "D6EAF8", BIRU),
        ("   Lempung: isi Cu (kPa), φ = 0", False, PUTIH, "000000"),
        ("   Pasir   : isi φ (°),  Cu = 0", False, PUTIH, "000000"),
        ("   Jika Cu & φ = 0: program estimasi otomatis dari SPT-N", False, PUTIH, "000000"),
        ("", False, PUTIH, "000000"),
        ("5. FORMAT FILE", True, "D6EAF8", BIRU),
        ("   Simpan sebagai .xlsx (Excel 2007 ke atas)", False, PUTIH, "000000"),
        ("   Google Sheets: File → Download → Microsoft Excel (.xlsx)", False, PUTIH, "000000"),
        ("   Format .xls (Excel 97-2003) juga diterima", False, PUTIH, "000000"),
        ("   Format .csv juga diterima (tanpa formatting)", False, PUTIH, "000000"),
        ("", False, PUTIH, "000000"),
        ("6. MENAMBAH LAPISAN", True, "D6EAF8", BIRU),
        ("   Klik nomor baris pada baris BATAS BAWAH → Insert Row Above", False, PUTIH, "000000"),
        ("   Isi data lapisan baru di baris yang baru ditambahkan", False, PUTIH, "000000"),
    ]
    ws2.column_dimensions["A"].width = 65
    for i, (teks, bold, bg, fc) in enumerate(petunjuk):
        cell = ws2.cell(i+1, 1, teks)
        cell.font = Font(name="Arial", bold=bold, size=10, color=fc)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = al_l
        ws2.row_dimensions[i+1].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



# ═══════════════════════════════════════════════════
# TABEL PERBANDINGAN METODE INPUT (untuk expander)
# ═══════════════════════════════════════════════════
_TABEL_PERBANDINGAN = """
| Metode | Keunggulan | Keterbatasan | Cocok untuk |
|--------|-----------|-------------|-------------|
| ✏️ Manual per lapisan | Paling fleksibel, parameter langsung dikontrol | Butuh interpretasi bore log terlebih dahulu | Punya data laboratorium / bore log yang sudah diinterpretasi |
| 📂 Upload Excel | Cepat untuk banyak lapisan, bisa disiapkan di kantor | Format harus sesuai template | Data sudah rapi di spreadsheet |
| 📊 SPT-N per kedalaman | Langsung dari data lapangan, otomatis layering | Parameter estimasi empiris (Peck 1974, Stroud 1974) | Punya data boring log mentah |
| 🔬 Sondir/CPT | Resolusi tinggi, klasifikasi SBT Robertson 1990 | Butuh data qc & fs, hasil sangat bergantung kualitas CPT | Punya data sondir terperinci |
| 🔀 SPT + Sondir (min) | Paling konservatif, dua sumber data | Butuh keduanya, proses lebih lama | Desain final / gedung penting |
| ⚡ Korelasi langsung | Paling cepat, tidak perlu data lapangan | Paling tidak akurat, hanya untuk preliminary | Estimasi awal / studi kelayakan |

**Saran program:** Untuk desain akhir gunakan data laboratorium atau CPT. SPT-N memadai untuk desain
bangunan sedang. Korelasi langsung **HANYA** untuk studi awal — jangan dipakai untuk konstruksi.
"""


def render_input_tanah_multi() -> pd.DataFrame:
    """
    Antarmuka input data tanah dengan 6 metode pilihan.
    Setiap metode disertai tooltip (help) dan tabel perbandingan di expander.
    """
    from utils.soil_converter import (
        spt_per_kedalaman_ke_lapisan,
        sondir_ke_lapisan,
        spt_plus_sondir_ke_lapisan,
        korelasi_langsung_ke_lapisan,
        PARAM_TIPIKAL,
    )

    st.subheader("Data Tanah")

    # ── Pilihan metode dengan tooltip ──────────────────────
    METODE_OPTS = {
        "✏️ Manual per lapisan":
            "Masukkan parameter setiap lapisan langsung ke tabel. "
            "Paling akurat jika Anda sudah memiliki interpretasi bore log.",
        "📂 Upload Excel":
            "Upload file Excel yang sudah diisi. Download form terlebih dahulu. "
            "Format harus sama persis dengan template.",
        "📊 SPT-N per kedalaman":
            "Input nilai SPT-N mentah per kedalaman pengujian. "
            "Program otomatis membagi lapisan dan mengestimasi parameter via korelasi empiris (Peck 1974).",
        "🔬 Sondir / CPT":
            "Input data sondir (qc dan fs) per kedalaman. "
            "Klasifikasi tanah via Robertson (1990). Resolusi lebih tinggi dari SPT.",
        "🔀 SPT + Sondir (konservatif)":
            "Gabungkan SPT dan sondir, ambil parameter yang lebih rendah (lebih aman). "
            "Disarankan untuk gedung penting atau desain final.",
        "⚡ Korelasi langsung (pilih jenis tanah)":
            "Pilih jenis tanah per lapisan, program mengisi parameter tipikal otomatis. "
            "HANYA untuk studi awal / preliminary — tidak untuk desain konstruksi.",
    }

    metode = st.radio(
        "Metode input data tanah:",
        options=list(METODE_OPTS.keys()),
        help="Pilih sesuai data yang tersedia. Hover di setiap opsi untuk keterangan singkat.",
    )

    # Tooltip per metode (ditampilkan sebagai info box)
    st.caption(f"ℹ️ {METODE_OPTS[metode]}")

    # ── Expander tabel perbandingan ────────────────────────
    with st.expander("📋 Perbandingan semua metode input — klik untuk lihat"):
        st.markdown(_TABEL_PERBANDINGAN)
        st.info(
            "**Urutan akurasi (terbaik → tercepat):** "
            "Lab + Bore Log → CPT/Sondir → SPT+Sondir → SPT per kedalaman → Manual → Korelasi langsung"
        )

    st.divider()
    df_tanah = None

    # ══════════════════════════════════════════════════════
    # METODE 1: MANUAL (fungsi lama, reuse)
    # ══════════════════════════════════════════════════════
    if "Manual" in metode:
        df_tanah = _render_manual()

    # ══════════════════════════════════════════════════════
    # METODE 2: UPLOAD EXCEL (fungsi lama, reuse)
    # ══════════════════════════════════════════════════════
    elif "Upload" in metode:
        df_tanah = _render_upload()

    # ══════════════════════════════════════════════════════
    # METODE 3: SPT PER KEDALAMAN
    # ══════════════════════════════════════════════════════
    elif "SPT-N per kedalaman" in metode:
        st.markdown("**Input SPT-N per Kedalaman Pengujian**")
        st.caption("Masukkan setiap titik pengujian SPT (per 1.5 m atau sesuai interval boring).")
        col_dl_spt, _ = st.columns([1,3])
        with col_dl_spt:
            st.download_button("⬇️ Download Template SPT Excel",
                data=_template_spt_excel(), file_name="template_spt_kedalaman.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tmpl_spt", use_container_width=True)

        col_spt1, col_spt2 = st.columns([1, 1])
        with col_spt1:
            n_titik = st.number_input("Jumlah titik SPT", min_value=2, max_value=50,
                                       value=8, step=1, key="n_titik_spt")
        with col_spt2:
            threshold = st.slider(
                "Sensitivitas pemisahan lapisan (%)",
                min_value=20, max_value=80, value=40, step=5,
                help="Persentase perubahan N-value yang memicu batas lapisan baru. "
                     "Nilai kecil = lebih banyak lapisan tipis. Nilai besar = lapisan lebih tebal.",
                key="threshold_spt"
            )

        # Tabel input SPT
        df_spt_init = pd.DataFrame({
            "Kedalaman (m)": [i * 1.5 for i in range(n_titik)],
            "SPT-N":         [4, 6, 8, 12, 18, 25, 35, 45][:n_titik] +
                              [40] * max(0, n_titik - 8)
        })
        if "df_spt_input" not in st.session_state:
            st.session_state.df_spt_input = df_spt_init

        if len(st.session_state.df_spt_input) != n_titik:
            st.session_state.df_spt_input = df_spt_init

        df_spt_edit = st.data_editor(
            st.session_state.df_spt_input,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "Kedalaman (m)": st.column_config.NumberColumn("Kedalaman (m)", min_value=0.0, step=0.5, format="%.2f"),
                "SPT-N":         st.column_config.NumberColumn("SPT-N", min_value=0, max_value=100, step=1),
            },
            hide_index=True, key="editor_spt_kedalaman"
        )
        st.session_state.df_spt_input = df_spt_edit

        # Konversi dan preview
        if st.button("🔄 Konversi SPT → Lapisan", key="btn_konversi_spt"):
            df_tanah = spt_per_kedalaman_ke_lapisan(df_spt_edit, threshold_persen=threshold)
            st.session_state["df_tanah_konversi"] = df_tanah
            st.success(f"Berhasil! {len(df_tanah)} lapisan terdeteksi.")

        if "df_tanah_konversi" in st.session_state and "SPT-N" in metode:
            df_tanah = st.session_state["df_tanah_konversi"]
            st.markdown("**Hasil konversi (bisa diedit):**")
            df_tanah = _render_hasil_konversi(df_tanah, key="edit_spt_hasil")

    # ══════════════════════════════════════════════════════
    # METODE 4: SONDIR / CPT
    # ══════════════════════════════════════════════════════
    elif "Sondir" in metode and "SPT" not in metode:
        st.markdown("**Input Data Sondir / CPT**")
        st.caption(
            "Masukkan nilai qc (tahanan ujung) dan fs (gesekan selimut) per kedalaman. "
            "Korelasi: Robertson (1990) Soil Behaviour Type."
        )
        col_dl_cpt, _ = st.columns([1,3])
        with col_dl_cpt:
            st.download_button("⬇️ Download Template Sondir Excel",
                data=_template_cpt_excel(), file_name="template_sondir_cpt.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tmpl_cpt", use_container_width=True)
        n_titik_cpt = st.number_input("Jumlah titik kedalaman", min_value=2, max_value=100,
                                       value=10, step=1, key="n_titik_cpt")
        df_cpt_init = pd.DataFrame({
            "Kedalaman (m)": [i * 0.2 * n_titik_cpt / 10 for i in range(n_titik_cpt)],
            "qc (MPa)":      [0.3, 0.5, 0.8, 1.2, 2.0, 3.5, 5.0, 8.0, 12.0, 15.0][:n_titik_cpt] +
                              [15.0] * max(0, n_titik_cpt - 10),
            "fs (kPa)":      [10, 15, 20, 30, 40, 60, 80, 100, 120, 130][:n_titik_cpt] +
                              [130] * max(0, n_titik_cpt - 10),
        })
        if "df_cpt_input" not in st.session_state:
            st.session_state.df_cpt_input = df_cpt_init

        df_cpt_edit = st.data_editor(
            st.session_state.df_cpt_input,
            use_container_width=True, num_rows="fixed",
            column_config={
                "Kedalaman (m)": st.column_config.NumberColumn(format="%.2f"),
                "qc (MPa)":      st.column_config.NumberColumn(format="%.2f", min_value=0.0),
                "fs (kPa)":      st.column_config.NumberColumn(format="%.1f", min_value=0.0),
            },
            hide_index=True, key="editor_cpt"
        )
        st.session_state.df_cpt_input = df_cpt_edit

        maw_cpt = st.number_input("Muka air tanah (m)", value=2.0, step=0.5, key="maw_cpt")
        if st.button("🔄 Konversi CPT → Lapisan", key="btn_konversi_cpt"):
            df_tanah = sondir_ke_lapisan(df_cpt_edit, maw_cpt)
            st.session_state["df_tanah_konversi_cpt"] = df_tanah
            st.success(f"Berhasil! {len(df_tanah)} lapisan terdeteksi.")

        if "df_tanah_konversi_cpt" in st.session_state and "Sondir" in metode and "SPT" not in metode:
            df_tanah = st.session_state["df_tanah_konversi_cpt"]
            st.markdown("**Hasil konversi (bisa diedit):**")
            df_tanah = _render_hasil_konversi(df_tanah, key="edit_cpt_hasil")

    # ══════════════════════════════════════════════════════
    # METODE 5: SPT + SONDIR
    # ══════════════════════════════════════════════════════
    elif "SPT + Sondir" in metode:
        st.markdown("**Input SPT-N + Sondir (akan diambil nilai minimum)**")
        st.info("Masukkan data SPT dan data Sondir secara terpisah. "
                "Program akan mengambil parameter yang lebih konservatif (lebih rendah).")

        col_m5a, col_m5b = st.columns(2)
        with col_m5a:
            st.markdown("*Data SPT-N*")
            n5 = st.number_input("Titik SPT", value=6, min_value=2, max_value=30, key="n_m5_spt")
            df_spt5 = pd.DataFrame({
                "Kedalaman (m)": [i*1.5 for i in range(n5)],
                "SPT-N": [5,8,12,20,30,40][:n5] + [40]*max(0,n5-6)
            })
            df_spt5_edit = st.data_editor(df_spt5, num_rows="fixed", hide_index=True,
                                           key="ed_m5_spt", use_container_width=True)

        with col_m5b:
            st.markdown("*Data Sondir (qc & fs)*")
            n5c = st.number_input("Titik CPT", value=8, min_value=2, max_value=50, key="n_m5_cpt")
            df_cpt5 = pd.DataFrame({
                "Kedalaman (m)": [i*0.8 for i in range(n5c)],
                "qc (MPa)":      [0.4,0.6,1.0,2.0,3.0,5.0,8.0,12.0][:n5c]+[12.0]*max(0,n5c-8),
                "fs (kPa)":      [12,18,25,40,55,75,95,110][:n5c]+[110]*max(0,n5c-8),
            })
            df_cpt5_edit = st.data_editor(df_cpt5, num_rows="fixed", hide_index=True,
                                           key="ed_m5_cpt", use_container_width=True)

        maw5 = st.number_input("Muka air tanah (m)", value=2.0, step=0.5, key="maw_m5")
        if st.button("🔄 Gabungkan & Konversi", key="btn_m5"):
            df_tanah = spt_plus_sondir_ke_lapisan(df_spt5_edit, df_cpt5_edit, maw5)
            st.session_state["df_tanah_m5"] = df_tanah
            st.success(f"Berhasil! {len(df_tanah)} lapisan (nilai konservatif).")

        if "df_tanah_m5" in st.session_state and "SPT + Sondir" in metode:
            df_tanah = st.session_state["df_tanah_m5"]
            st.markdown("**Hasil gabungan (bisa diedit):**")
            df_tanah = _render_hasil_konversi(df_tanah, key="edit_m5_hasil")

    # ══════════════════════════════════════════════════════
    # METODE 6: KORELASI LANGSUNG
    # ══════════════════════════════════════════════════════
    elif "Korelasi" in metode:
        st.warning(
            "⚠️ **Hanya untuk studi awal / preliminary design.** "
            "Parameter yang digunakan adalah nilai tipikal — BUKAN data aktual lapangan. "
            "Jangan gunakan untuk desain konstruksi final."
        )
        n_lap_kor = st.number_input("Jumlah lapisan", min_value=1, max_value=15, value=4, key="n_kor")

        lapisan_kor = []
        z_prev = 0.0
        for i in range(int(n_lap_kor)):
            with st.expander(f"Lapisan {i+1}", expanded=(i == 0)):
                col_k1, col_k2, col_k3 = st.columns(3)
                with col_k1:
                    jenis_k = st.selectbox(
                        "Jenis tanah", JENIS_TANAH_OPSI,
                        key=f"kor_jenis_{i}",
                        help=f"Program mengisi Cu, φ, γ otomatis dari pilihan ini."
                    )
                with col_k2:
                    z_atas_k = st.number_input("z atas (m)", value=float(z_prev),
                                                key=f"kor_zatas_{i}", step=0.5, format="%.1f")
                with col_k3:
                    z_bawah_k = st.number_input("z bawah (m)", value=float(z_prev + 3.0),
                                                 key=f"kor_zbawah_{i}", step=0.5, format="%.1f")
                p_tip = PARAM_TIPIKAL.get(jenis_k, {})
                st.caption(
                    f"Parameter tipikal: Cu={p_tip.get('cu',0):.0f} kPa · "
                    f"φ={p_tip.get('phi',0):.0f}° · γ={p_tip.get('gamma',17):.0f} kN/m³ · "
                    f"SPT-N≈{p_tip.get('spt',10)}"
                )
                z_prev = z_bawah_k
                lapisan_kor.append({"jenis": jenis_k, "z_atas": z_atas_k, "z_bawah": z_bawah_k})

        if st.button("✅ Gunakan parameter tipikal ini", key="btn_kor"):
            df_tanah = korelasi_langsung_ke_lapisan(lapisan_kor)
            st.session_state["df_tanah_kor"] = df_tanah
            st.success("Parameter tipikal diterapkan. Periksa hasilnya di bawah.")

        if "df_tanah_kor" in st.session_state and "Korelasi" in metode:
            df_tanah = st.session_state["df_tanah_kor"]
            st.markdown("**Data tanah dari korelasi (bisa diedit):**")
            df_tanah = _render_hasil_konversi(df_tanah, key="edit_kor_hasil")

    return df_tanah if df_tanah is not None else DATA_TANAH_DEFAULT.copy()



def _render_upload() -> pd.DataFrame:
    """Render upload Excel dengan download template dan deteksi format otomatis."""
    st.caption("Upload file Excel yang sudah diisi. Download form terlebih dahulu.")

    col_tmpl, col_info = st.columns([1, 2])
    with col_tmpl:
        buf_tmpl = _buat_template_excel_lengkap()
        st.download_button(
            label="⬇️ Download Form Excel",
            data=buf_tmpl,
            file_name="form_input_data_tanah.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download form Excel yang sudah terformat. Isi data, lalu upload kembali.",
            use_container_width=True,
        )
    with col_info:
        st.caption(
            "Form berisi contoh data, petunjuk pengisian, dan batas kolom/baris. "
            "Format: .xlsx, .xls, atau .csv. Google Sheets: File → Download → Excel (.xlsx)."
        )

    file_upload = st.file_uploader(
        "Upload file Excel data tanah:",
        type=["xlsx", "xls", "csv"],
        help="Upload form yang sudah diisi."
    )

    if file_upload is not None:
        try:
            if file_upload.name.endswith('.csv'):
                df_upload = pd.read_csv(file_upload)
            else:
                df_cek = pd.read_excel(file_upload, nrows=3)
                file_upload.seek(0)
                kolom_standar = KOLOM_TANAH["z_atas"]
                if kolom_standar not in df_cek.columns:
                    df_upload = pd.read_excel(file_upload, header=2, skiprows=[3])
                else:
                    df_upload = pd.read_excel(file_upload)

            # Filter baris numerik (hapus baris batas teks merah)
            if KOLOM_TANAH["z_atas"] in df_upload.columns:
                df_upload = df_upload[
                    pd.to_numeric(df_upload[KOLOM_TANAH["z_atas"]], errors="coerce").notna()
                ].copy().reset_index(drop=True)

            kolom_wajib = [KOLOM_TANAH["z_atas"], KOLOM_TANAH["z_bawah"],
                           KOLOM_TANAH["jenis"], KOLOM_TANAH["spt"]]
            kolom_hilang = [k for k in kolom_wajib if k not in df_upload.columns]
            if kolom_hilang:
                st.error(f"Kolom tidak ditemukan: {kolom_hilang}. Gunakan template di atas.")
                return DATA_TANAH_DEFAULT.copy()

            for k, default in [(KOLOM_TANAH["cu"], 0.0), (KOLOM_TANAH["phi"], 0.0),
                                (KOLOM_TANAH["gamma"], 17.0)]:
                if k not in df_upload.columns:
                    df_upload[k] = default
            df_upload[KOLOM_TANAH["no"]] = range(1, len(df_upload) + 1)

            df_tanah = df_upload[list(KOLOM_TANAH.values())].copy()
            st.success(f"✅ File berhasil dibaca: {len(df_tanah)} lapisan tanah.")
            st.dataframe(df_tanah, use_container_width=True, hide_index=True)
            return df_tanah

        except Exception as e:
            st.error(f"Gagal membaca file: {e}")
            return DATA_TANAH_DEFAULT.copy()
    else:
        st.info("Belum ada file diupload. Gunakan template di atas, isi data, lalu upload.")
        return DATA_TANAH_DEFAULT.copy()


def _template_spt_excel() -> bytes:
    """Template Excel untuk input SPT-N per kedalaman."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "SPT_Kedalaman"
    BIRU = "1A5376"; KNG = "FFF9C4"; f_h = Font(name="Arial",bold=True,color="FFFFFF",size=10)
    f_b = Font(name="Arial",size=10)
    fh = PatternFill("solid",fgColor=BIRU); fk = PatternFill("solid",fgColor=KNG)
    s = Side(style="thin",color="AAAAAA"); brd = Border(left=s,right=s,top=s,bottom=s)
    ac = Alignment(horizontal="center",vertical="center")
    ws.merge_cells("A1:B1")
    ws["A1"].value = "FORM INPUT SPT-N PER KEDALAMAN — Program Pile Capacity"
    ws["A1"].font = f_h; ws["A1"].fill = fh; ws["A1"].alignment = ac
    ws.row_dimensions[1].height = 22
    for i,hdr in enumerate(["Kedalaman (m)","SPT-N"]):
        c = ws.cell(2,i+1,hdr); c.font=f_h; c.fill=fh; c.alignment=ac; c.border=brd
    contoh = [(0.0,4),(1.5,6),(3.0,9),(4.5,12),(6.0,18),(7.5,25),(9.0,35),(10.5,45)]
    for i,(z,n) in enumerate(contoh):
        for j,val in enumerate([z,n]):
            c = ws.cell(3+i,j+1,val); c.font=f_b; c.fill=fk; c.alignment=ac; c.border=brd
    ws.column_dimensions["A"].width=18; ws.column_dimensions["B"].width=14
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _template_cpt_excel() -> bytes:
    """Template Excel untuk input Sondir/CPT."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Sondir_CPT"
    BIRU = "1A5376"; KNG = "FFF9C4"; f_h = Font(name="Arial",bold=True,color="FFFFFF",size=10)
    f_b = Font(name="Arial",size=10)
    fh = PatternFill("solid",fgColor=BIRU); fk = PatternFill("solid",fgColor=KNG)
    s = Side(style="thin",color="AAAAAA"); brd = Border(left=s,right=s,top=s,bottom=s)
    ac = Alignment(horizontal="center",vertical="center")
    ws.merge_cells("A1:C1")
    ws["A1"].value = "FORM INPUT SONDIR/CPT — Program Pile Capacity"
    ws["A1"].font = f_h; ws["A1"].fill = fh; ws["A1"].alignment = ac
    ws.row_dimensions[1].height = 22
    for i,hdr in enumerate(["Kedalaman (m)","qc (MPa)","fs (kPa)"]):
        c = ws.cell(2,i+1,hdr); c.font=f_h; c.fill=fh; c.alignment=ac; c.border=brd
    contoh = [(0.0,0.3,10),(0.5,0.5,15),(1.0,0.8,20),(2.0,1.5,35),
              (3.0,2.5,50),(5.0,4.0,70),(8.0,6.0,90),(12.0,9.0,110),(15.0,12.0,120),(20.0,15.0,130)]
    for i,row in enumerate(contoh):
        for j,val in enumerate(row):
            c = ws.cell(3+i,j+1,val); c.font=f_b; c.fill=fk; c.alignment=ac; c.border=brd
    for col,w in zip(["A","B","C"],[16,14,14]): ws.column_dimensions[col].width=w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _render_hasil_konversi(df: pd.DataFrame, key: str) -> pd.DataFrame:
    """Tampilkan hasil konversi dalam tabel yang bisa diedit."""
    return st.data_editor(
        df, use_container_width=True, num_rows="fixed",
        column_config={
            KOLOM_TANAH["no"]:     st.column_config.NumberColumn("No.", disabled=True, width="small"),
            KOLOM_TANAH["z_atas"]: st.column_config.NumberColumn("z atas (m)", format="%.2f"),
            KOLOM_TANAH["z_bawah"]:st.column_config.NumberColumn("z bawah (m)", format="%.2f"),
            KOLOM_TANAH["jenis"]:  st.column_config.SelectboxColumn("Jenis Tanah", options=JENIS_TANAH_OPSI),
            KOLOM_TANAH["spt"]:    st.column_config.NumberColumn("SPT-N", min_value=0, max_value=100),
            KOLOM_TANAH["cu"]:     st.column_config.NumberColumn("Cu (kPa)", format="%.1f"),
            KOLOM_TANAH["phi"]:    st.column_config.NumberColumn("φ (°)", format="%.1f"),
            KOLOM_TANAH["gamma"]:  st.column_config.NumberColumn("γ (kN/m³)", format="%.1f"),
        },
        hide_index=True, key=key
    )


def _render_manual() -> pd.DataFrame:
    """Render input manual (fungsi lama, dipindah ke sini)."""
    st.caption("Edit tabel di bawah langsung — klik sel untuk mengubah nilai.")
    if "df_tanah_manual" not in st.session_state:
        st.session_state.df_tanah_manual = DATA_TANAH_DEFAULT.copy()

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("➕ Tambah lapisan", use_container_width=True):
            b = pd.DataFrame({
                KOLOM_TANAH["no"]:     [len(st.session_state.df_tanah_manual)+1],
                KOLOM_TANAH["z_atas"]: [st.session_state.df_tanah_manual[KOLOM_TANAH["z_bawah"]].iloc[-1]],
                KOLOM_TANAH["z_bawah"]:[st.session_state.df_tanah_manual[KOLOM_TANAH["z_bawah"]].iloc[-1]+3.0],
                KOLOM_TANAH["jenis"]:  ["Lempung sedang"],
                KOLOM_TANAH["spt"]:    [10], KOLOM_TANAH["cu"]: [30.0],
                KOLOM_TANAH["phi"]:    [0.0], KOLOM_TANAH["gamma"]: [17.0],
            })
            st.session_state.df_tanah_manual = pd.concat(
                [st.session_state.df_tanah_manual, b], ignore_index=True)
    with col2:
        if st.button("➖ Hapus lapisan terakhir", use_container_width=True):
            if len(st.session_state.df_tanah_manual) > 1:
                st.session_state.df_tanah_manual = st.session_state.df_tanah_manual.iloc[:-1].copy()
    with col3:
        if st.button("🔄 Reset ke default", use_container_width=True):
            st.session_state.df_tanah_manual = DATA_TANAH_DEFAULT.copy()

    df_edit = st.data_editor(
        st.session_state.df_tanah_manual, use_container_width=True,
        num_rows="fixed",
        column_config={
            KOLOM_TANAH["no"]:     st.column_config.NumberColumn("No.", disabled=True, width="small"),
            KOLOM_TANAH["z_atas"]: st.column_config.NumberColumn("z atas (m)", min_value=0.0, step=0.5, format="%.1f"),
            KOLOM_TANAH["z_bawah"]:st.column_config.NumberColumn("z bawah (m)", min_value=0.1, step=0.5, format="%.1f"),
            KOLOM_TANAH["jenis"]:  st.column_config.SelectboxColumn("Jenis Tanah", options=JENIS_TANAH_OPSI),
            KOLOM_TANAH["spt"]:    st.column_config.NumberColumn("SPT-N", min_value=0, max_value=100, step=1),
            KOLOM_TANAH["cu"]:     st.column_config.NumberColumn("Cu (kPa)", min_value=0.0, step=5.0, format="%.1f"),
            KOLOM_TANAH["phi"]:    st.column_config.NumberColumn("φ (°)", min_value=0.0, max_value=45.0, format="%.1f"),
            KOLOM_TANAH["gamma"]:  st.column_config.NumberColumn("γ (kN/m³)", min_value=10.0, max_value=25.0, format="%.1f"),
        },
        hide_index=True, key="editor_tanah"
    )
    st.session_state.df_tanah_manual = df_edit.copy()
    return df_edit.copy()

def render_input_tanah() -> pd.DataFrame:
    """
    Menampilkan antarmuka input data tanah.
    User bisa memilih: input manual via tabel, atau upload file Excel.
    Mengembalikan DataFrame data tanah yang sudah divalidasi.
    """
    st.subheader("Data Tanah")

    cara_input = st.radio(
        "Cara input data tanah:",
        options=["✏️ Input manual (edit tabel langsung)", "📂 Upload file Excel (.xlsx)"],
        horizontal=True,
        help="Pilih cara yang paling nyaman. Format kolom harus sama untuk kedua cara."
    )

    df_tanah = None

    # ----------------------------------------------------------
    # CARA 1: Input manual via st.data_editor
    # ----------------------------------------------------------
    if "manual" in cara_input:
        st.caption("Edit tabel di bawah langsung — klik sel untuk mengubah nilai.")

        # Inisialisasi session state agar tabel tidak reset saat widget lain berubah
        if "df_tanah_manual" not in st.session_state:
            st.session_state.df_tanah_manual = DATA_TANAH_DEFAULT.copy()

        col_tambah, col_hapus, col_reset = st.columns([1, 1, 1])
        with col_tambah:
            if st.button("➕ Tambah lapisan", use_container_width=True):
                baris_baru = pd.DataFrame({
                    KOLOM_TANAH["no"]:     [len(st.session_state.df_tanah_manual) + 1],
                    KOLOM_TANAH["z_atas"]: [st.session_state.df_tanah_manual[KOLOM_TANAH["z_bawah"]].iloc[-1]],
                    KOLOM_TANAH["z_bawah"]:[st.session_state.df_tanah_manual[KOLOM_TANAH["z_bawah"]].iloc[-1] + 3.0],
                    KOLOM_TANAH["jenis"]:  ["Lempung sedang"],
                    KOLOM_TANAH["spt"]:    [10],
                    KOLOM_TANAH["cu"]:     [30.0],
                    KOLOM_TANAH["phi"]:    [0.0],
                    KOLOM_TANAH["gamma"]:  [17.0],
                })
                st.session_state.df_tanah_manual = pd.concat(
                    [st.session_state.df_tanah_manual, baris_baru], ignore_index=True
                )
        with col_hapus:
            if st.button("➖ Hapus lapisan terakhir", use_container_width=True):
                if len(st.session_state.df_tanah_manual) > 1:
                    st.session_state.df_tanah_manual = st.session_state.df_tanah_manual.iloc[:-1].copy()
        with col_reset:
            if st.button("🔄 Reset ke default", use_container_width=True):
                st.session_state.df_tanah_manual = DATA_TANAH_DEFAULT.copy()

        df_edit = st.data_editor(
            st.session_state.df_tanah_manual,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                KOLOM_TANAH["no"]:     st.column_config.NumberColumn("No.", disabled=True, width="small"),
                KOLOM_TANAH["z_atas"]: st.column_config.NumberColumn("z atas (m)", min_value=0.0, step=0.5, format="%.1f"),
                KOLOM_TANAH["z_bawah"]:st.column_config.NumberColumn("z bawah (m)", min_value=0.1, step=0.5, format="%.1f"),
                KOLOM_TANAH["jenis"]:  st.column_config.SelectboxColumn("Jenis Tanah", options=JENIS_TANAH_OPSI),
                KOLOM_TANAH["spt"]:    st.column_config.NumberColumn("SPT-N", min_value=0, max_value=100, step=1),
                KOLOM_TANAH["cu"]:     st.column_config.NumberColumn("Cu (kPa)", min_value=0.0, step=5.0, format="%.1f",
                                           help="Isi 0 jika tanah pasir/non-kohesif"),
                KOLOM_TANAH["phi"]:    st.column_config.NumberColumn("φ (°)", min_value=0.0, max_value=45.0, step=1.0, format="%.1f",
                                           help="Isi 0 jika tanah lempung/kohesif"),
                KOLOM_TANAH["gamma"]:  st.column_config.NumberColumn("γ (kN/m³)", min_value=10.0, max_value=25.0, step=0.5, format="%.1f"),
            },
            hide_index=True,
            key="editor_tanah",
        )
        st.session_state.df_tanah_manual = df_edit.copy()
        df_tanah = df_edit.copy()

    # ----------------------------------------------------------
    # CARA 2: Upload file Excel
    # ----------------------------------------------------------
    else:
        st.caption("Upload file Excel dengan kolom sesuai template di bawah.")

        # Tombol download template — versi lengkap dengan keterangan
        col_tmpl, col_info = st.columns([1, 2])
        with col_tmpl:
            buf_tmpl = _buat_template_excel_lengkap()
            st.download_button(
                label="⬇️ Download Form Excel",
                data=buf_tmpl,
                file_name="form_input_data_tanah.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download form Excel yang sudah terformat. Isi data, lalu upload kembali.",
                use_container_width=True,
            )
        with col_info:
            st.caption(
                "Form berisi contoh data, petunjuk pengisian, dan batas kolom/baris. "
                "Format: .xlsx (Excel 2007+). Google Sheets: File → Download → Excel (.xlsx)."
            )

        file_upload = st.file_uploader(
            "Upload file Excel data tanah:",
            type=["xlsx", "xls", "csv"],
            help="Upload form yang sudah diisi. Format: .xlsx, .xls, atau .csv."
        )

        if file_upload is not None:
            try:
                if file_upload.name.endswith('.csv'):
                    df_upload = pd.read_csv(file_upload)
                else:
                    # Coba baca dulu tanpa skip untuk deteksi format
                    df_cek = pd.read_excel(file_upload, nrows=3)
                    file_upload.seek(0)
                    # Jika kolom pertama bukan nama kolom standar (ada judul di baris 1-2),
                    # skip 2 baris pertama lalu skip lagi baris satuan (baris ke-4 = index 1 setelah header)
                    kolom_standar = KOLOM_TANAH["z_atas"]
                    if kolom_standar not in df_cek.columns:
                        # Format template lengkap: header di baris 3 (skiprows=2), ada baris satuan
                        df_upload = pd.read_excel(file_upload, header=2, skiprows=[3])
                        file_upload.seek(0)
                    else:
                        df_upload = pd.read_excel(file_upload)
                # Hapus baris yang kosong atau berisi teks batas
                if KOLOM_TANAH["z_atas"] in df_upload.columns:
                    # Filter: hanya baris yang nilai z_atas-nya numerik
                    df_upload = df_upload[
                        pd.to_numeric(df_upload[KOLOM_TANAH["z_atas"]], errors='coerce').notna()
                    ].copy()
                    df_upload = df_upload.reset_index(drop=True)

                # Cek kolom minimum yang wajib ada
                kolom_wajib = [KOLOM_TANAH["z_atas"], KOLOM_TANAH["z_bawah"],
                               KOLOM_TANAH["jenis"], KOLOM_TANAH["spt"]]
                kolom_hilang = [k for k in kolom_wajib if k not in df_upload.columns]
                if kolom_hilang:
                    st.error(f"Kolom berikut tidak ditemukan di file: {kolom_hilang}")
                    st.info("Pastikan nama kolom sama persis dengan template.")
                else:
                    # Isi kolom opsional yang mungkin tidak ada
                    for k, default in [(KOLOM_TANAH["cu"], 0.0),
                                       (KOLOM_TANAH["phi"], 0.0),
                                       (KOLOM_TANAH["gamma"], 17.0),
                                       (KOLOM_TANAH["no"], None)]:
                        if k not in df_upload.columns:
                            df_upload[k] = default
                    if KOLOM_TANAH["no"] not in df_upload.columns or df_upload[KOLOM_TANAH["no"]].isna().all():
                        df_upload[KOLOM_TANAH["no"]] = range(1, len(df_upload) + 1)

                    df_tanah = df_upload[list(KOLOM_TANAH.values())].copy()
                    st.success(f"File berhasil dibaca: {len(df_tanah)} lapisan tanah.")
                    st.dataframe(df_tanah, use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
        else:
            st.info("Belum ada file yang diupload. Gunakan template di atas.")
            df_tanah = DATA_TANAH_DEFAULT.copy()

    return df_tanah


def validasi_data_tanah(df: pd.DataFrame) -> tuple[bool, list[str]]:
    """
    Memvalidasi DataFrame data tanah.
    Mengembalikan (valid: bool, pesan_error: list[str])
    """
    pesan = []

    if df is None or len(df) == 0:
        return False, ["Data tanah kosong. Masukkan minimal 1 lapisan tanah."]

    z_atas  = df[KOLOM_TANAH["z_atas"]].values
    z_bawah = df[KOLOM_TANAH["z_bawah"]].values

    # Cek kedalaman bawah > kedalaman atas
    for i, (za, zb) in enumerate(zip(z_atas, z_bawah)):
        if zb <= za:
            pesan.append(f"Lapisan {i+1}: kedalaman bawah ({zb} m) harus lebih besar dari kedalaman atas ({za} m).")

    # Cek kontinuitas antar lapisan (bawah lapisan i = atas lapisan i+1)
    for i in range(len(df) - 1):
        if abs(z_bawah[i] - z_atas[i+1]) > 0.01:
            pesan.append(
                f"Lapisan {i+1} → {i+2}: ada jeda/tumpang tindih kedalaman "
                f"({z_bawah[i]:.1f} m ≠ {z_atas[i+1]:.1f} m). "
                "Kedalaman bawah lapisan atas harus sama dengan kedalaman atas lapisan bawah."
            )

    # Cek SPT-N tidak negatif
    spt = df[KOLOM_TANAH["spt"]].values
    if any(s < 0 for s in spt):
        pesan.append("Nilai SPT-N tidak boleh negatif.")

    # Cek gamma masuk akal
    gamma = df[KOLOM_TANAH["gamma"]].values
    for i, g in enumerate(gamma):
        if g < 10.0 or g > 25.0:
            pesan.append(f"Lapisan {i+1}: γ = {g} kN/m³ di luar rentang wajar (10–25 kN/m³).")

    valid = len(pesan) == 0
    return valid, pesan


def render_input_tiang() -> dict:
    """
    Menampilkan input parameter tiang di sidebar.
    Mengembalikan dict parameter tiang.
    """
    st.sidebar.header("Parameter Tiang")

    tipe_tiang = st.sidebar.selectbox(
        "Tipe tiang",
        options=["Spun pile (bulat, pracetak)", "Square pile (kotak, pracetak)",
                 "Boredpile (beton cor, bulat)", "H-pile baja"],
        help="Tipe tiang mempengaruhi koefisien α/β dan metode end bearing."
    )

    # Dimensi tiang tergantung tipe
    if "Spun" in tipe_tiang or "Boredpile" in tipe_tiang:
        diameter = st.sidebar.number_input(
            "Diameter tiang (m)", min_value=0.10, max_value=2.50,
            value=0.40, step=0.05, format="%.2f"
        )
        lebar = None  # tidak dipakai untuk tiang bulat
    elif "Square" in tipe_tiang:
        lebar = st.sidebar.number_input(
            "Lebar sisi tiang (m)", min_value=0.10, max_value=1.50,
            value=0.35, step=0.05, format="%.2f"
        )
        diameter = None
    else:  # H-pile
        st.sidebar.caption("Dimensi H-pile (profil standar)")
        lebar_sayap = st.sidebar.number_input("Lebar sayap bf (m)", value=0.200, step=0.010, format="%.3f")
        tinggi_profil = st.sidebar.number_input("Tinggi profil d (m)", value=0.200, step=0.010, format="%.3f")
        tebal_badan = st.sidebar.number_input("Tebal badan tw (m)", value=0.009, step=0.001, format="%.3f")
        tebal_sayap = st.sidebar.number_input("Tebal sayap tf (m)", value=0.014, step=0.001, format="%.3f")
        diameter = None
        lebar = lebar_sayap

    kedalaman_tiang = st.sidebar.number_input(
        "Kedalaman tiang (m)", min_value=1.0, max_value=80.0,
        value=20.0, step=1.0, format="%.1f",
        help="Kedalaman dari muka tanah ke ujung bawah tiang."
    )

    variasi_kedalaman = st.sidebar.checkbox(
        "Hitung variasi kedalaman (untuk grafik Qtotal vs L)",
        value=True,
        help="Jika dicentang, program menghitung kapasitas untuk berbagai kedalaman tiang."
    )

    if variasi_kedalaman:
        col1, col2 = st.sidebar.columns(2)
        with col1:
            z_min = st.number_input("L min (m)", value=5.0, step=1.0, key="z_min_var")
        with col2:
            z_max = st.number_input("L maks (m)", value=kedalaman_tiang, step=1.0, key="z_max_var")
    else:
        z_min = kedalaman_tiang
        z_max = kedalaman_tiang

    # Material beton (tidak berlaku untuk H-pile)
    if "H-pile" not in tipe_tiang:
        material = st.sidebar.selectbox(
            "Mutu beton",
            options=["K-400 (fc' = 33.2 MPa)", "K-500 (fc' = 41.5 MPa)"],
        )
        fc_prime = 33.2 if "K-400" in material else 41.5
    else:
        material = "Baja A36 (fy = 250 MPa)"
        fc_prime = None

    muka_air = st.sidebar.number_input(
        "Kedalaman muka air tanah (m)", min_value=0.0, max_value=50.0,
        value=2.0, step=0.5, format="%.1f",
        help="Digunakan untuk menghitung tegangan efektif σ'v."
    )

    sf_tekan = st.sidebar.number_input(
        "Safety factor (tekan)", value=3.0, min_value=1.5, max_value=5.0,
        step=0.5, format="%.1f"
    )
    sf_tarik = st.sidebar.number_input(
        "Safety factor (tarik)", value=2.5, min_value=1.5, max_value=5.0,
        step=0.5, format="%.1f"
    )

    # Hitung properti geometri tiang
    if "Spun" in tipe_tiang or "Boredpile" in tipe_tiang:
        area_ujung  = np.pi / 4 * diameter**2
        keliling    = np.pi * diameter
        dim_label   = f"D = {diameter*100:.0f} cm"
    elif "Square" in tipe_tiang:
        area_ujung  = lebar**2
        keliling    = 4 * lebar
        diameter    = lebar  # gunakan sebagai referensi
        dim_label   = f"b = {lebar*100:.0f} cm"
    else:  # H-pile
        # Keliling H-pile (perimeter box untuk gesekan selimut)
        keliling    = 2 * (lebar_sayap + tinggi_profil)
        area_badan  = tinggi_profil * tebal_badan
        area_sayap  = 2 * lebar_sayap * tebal_sayap
        area_profil = area_badan + area_sayap
        area_ujung  = lebar_sayap * tinggi_profil  # plug area
        diameter    = (lebar_sayap + tinggi_profil) / 2  # ekuivalen
        dim_label   = f"bf={lebar_sayap*100:.0f}cm × d={tinggi_profil*100:.0f}cm"

    return {
        "tipe":             tipe_tiang,
        "diameter":         diameter,
        "lebar":            lebar if lebar else diameter,
        "area_ujung":       area_ujung,
        "keliling":         keliling,
        "kedalaman":        kedalaman_tiang,
        "variasi_kedalaman":variasi_kedalaman,
        "z_min":            z_min if variasi_kedalaman else kedalaman_tiang,
        "z_max":            z_max if variasi_kedalaman else kedalaman_tiang,
        "material":         material,
        "fc_prime":         fc_prime,
        "muka_air":         muka_air,
        "sf_tekan":         sf_tekan,
        "sf_tarik":         sf_tarik,
        "dim_label":        dim_label,
        "is_displacement":  "Spun" in tipe_tiang or "Square" in tipe_tiang,
        "is_bored":         "Boredpile" in tipe_tiang,
        "is_hpile":         "H-pile" in tipe_tiang,
    }
