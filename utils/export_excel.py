# utils/export_excel.py
# Export hasil perhitungan ke Excel (.xlsx)
# PENTING: Hanya menulis VALUES (nilai), TIDAK ADA FORMULA
# Tujuan: output komersial — user hanya bisa melihat angka, bukan cara hitung
#
# Sheet layout:
#   Sheet 1: Info & Input       — data proyek dan parameter tiang
#   Sheet 2: Profil Tanah       — data SPT per lapisan
#   Sheet 3: Hasil Tekan_Tarik  — distribusi skin friction + ringkasan kapasitas
#   Sheet 4: Gaya Lateral       — hasil Broms atau p-y curve
#   Sheet 5: Grafik             — grafik embed sebagai gambar

import io
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ==============================================================
# KONSTANTA STYLE
# ==============================================================
# Warna (hex tanpa #)
BIRU_TUA   = "1A5376"   # header utama
BIRU_MUDA  = "D6EAF8"   # baris genap / sub-header
HIJAU_TUA  = "1E8449"   # angka positif / OK
MERAH_TUA  = "922B21"   # angka negatif / TIDAK OK
ABU_TERANG = "F2F3F4"   # baris ganjil
PUTIH      = "FFFFFF"
KUNING     = "FEF9E7"   # highlight ringkasan

FONT_HEADER = Font(name="Arial", bold=True, color=PUTIH, size=11)
FONT_SUBHDR = Font(name="Arial", bold=True, color=BIRU_TUA, size=10)
FONT_BODY   = Font(name="Arial", size=10)
FONT_ANGKA  = Font(name="Arial", size=10)
FONT_JUDUL  = Font(name="Arial", bold=True, size=14, color=BIRU_TUA)
FONT_KECIL  = Font(name="Arial", size=9, italic=True, color="555555")

FILL_HEADER = PatternFill("solid", fgColor=BIRU_TUA)
FILL_SUBHDR = PatternFill("solid", fgColor=BIRU_MUDA)
FILL_GENAP  = PatternFill("solid", fgColor=BIRU_MUDA)
FILL_GANJIL = PatternFill("solid", fgColor=ABU_TERANG)
FILL_KUNING = PatternFill("solid", fgColor=KUNING)
FILL_PUTIH  = PatternFill("solid", fgColor=PUTIH)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

SISI_TIPIS = Side(style="thin",   color="BBBBBB")
SISI_TEBAL = Side(style="medium", color=BIRU_TUA)
BORDER_ALL = Border(left=SISI_TIPIS, right=SISI_TIPIS,
                     top=SISI_TIPIS,  bottom=SISI_TIPIS)
BORDER_TOP = Border(top=SISI_TEBAL)

FMT_ANGKA2  = '#,##0.00'      # 2 desimal dengan ribuan
FMT_ANGKA3  = '#,##0.000'     # 3 desimal
FMT_ANGKA0  = '#,##0'         # bulat
FMT_PERSEN  = '0.00%'


# ==============================================================
# HELPER FUNGSI STYLE
# ==============================================================

def _style_sel(ws, baris: int, kolom: int,
               nilai=None, font=None, fill=None,
               align=None, border=None, fmt: str = None,
               merge_ke: int = None) -> None:
    """Mengisi dan memberi style satu sel."""
    sel = ws.cell(row=baris, column=kolom)
    if nilai is not None:
        sel.value = nilai
    if font   is not None: sel.font      = font
    if fill   is not None: sel.fill      = fill
    if align  is not None: sel.alignment = align
    if border is not None: sel.border    = border
    if fmt    is not None: sel.number_format = fmt
    if merge_ke is not None:
        ws.merge_cells(
            start_row=baris, start_column=kolom,
            end_row=baris,   end_column=merge_ke
        )


def _baris_header(ws, baris: int, kolom_mulai: int,
                  judul_kolom: list[str]) -> None:
    """Membuat baris header tabel dengan warna biru tua."""
    for i, judul in enumerate(judul_kolom):
        _style_sel(ws, baris, kolom_mulai + i,
                   nilai=judul,
                   font=FONT_HEADER,
                   fill=FILL_HEADER,
                   align=ALIGN_CENTER,
                   border=BORDER_ALL)


def _baris_data(ws, baris: int, kolom_mulai: int,
                data: list, rata_kanan: list[int] = None,
                nomor_baris: int = 0) -> None:
    """Mengisi baris data dengan warna alternating."""
    if rata_kanan is None:
        rata_kanan = []
    fill = FILL_GENAP if nomor_baris % 2 == 0 else FILL_GANJIL

    for i, nilai in enumerate(data):
        align = ALIGN_RIGHT if i in rata_kanan else ALIGN_LEFT
        # Format angka otomatis
        fmt = None
        if isinstance(nilai, float):
            fmt = FMT_ANGKA2
        elif isinstance(nilai, int) and i in rata_kanan:
            fmt = FMT_ANGKA0
        _style_sel(ws, baris, kolom_mulai + i,
                   nilai=nilai, font=FONT_BODY,
                   fill=fill, align=align,
                   border=BORDER_ALL, fmt=fmt)


def _set_lebar_kolom(ws, lebar: dict) -> None:
    """Mengatur lebar kolom. lebar = {huruf: lebar_karakter}"""
    for huruf, w in lebar.items():
        ws.column_dimensions[huruf].width = w


def _tambah_judul_sheet(ws, teks: str, baris: int = 1,
                         kolom_mulai: int = 1, kolom_akhir: int = 8) -> None:
    """Menambahkan judul sheet yang di-merge."""
    _style_sel(ws, baris, kolom_mulai, nilai=teks,
               font=FONT_JUDUL, fill=FILL_PUTIH,
               align=ALIGN_LEFT, merge_ke=kolom_akhir)


def _tambah_info_baris(ws, baris: int, label: str, nilai,
                        kolom_label: int = 1, kolom_nilai: int = 2) -> None:
    """Satu baris label-nilai untuk tabel info."""
    _style_sel(ws, baris, kolom_label, nilai=label,
               font=FONT_SUBHDR, fill=FILL_SUBHDR,
               align=ALIGN_LEFT, border=BORDER_ALL)
    _style_sel(ws, baris, kolom_nilai, nilai=nilai,
               font=FONT_BODY, fill=FILL_PUTIH,
               align=ALIGN_LEFT, border=BORDER_ALL)


def _gambar_ke_bytes(fig: plt.Figure, dpi: int = 120) -> io.BytesIO:
    """Konversi Figure matplotlib ke BytesIO PNG."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


# ==============================================================
# SHEET 1: INFO & INPUT
# ==============================================================

def _buat_sheet_info(ws, param_tiang: dict,
                     nama_proyek: str, nama_konsultan: str,
                     nomor_laporan: str) -> None:
    ws.title = "1_Info_Input"
    _set_lebar_kolom(ws, {"A": 30, "B": 25, "C": 15, "D": 25, "E": 15})

    _tambah_judul_sheet(ws, "LAPORAN PERHITUNGAN KAPASITAS PONDASI TIANG", 1, 1, 5)
    ws.row_dimensions[1].height = 28

    # Blok info proyek
    _style_sel(ws, 2, 1, nilai="INFORMASI PROYEK",
               font=FONT_HEADER, fill=FILL_HEADER,
               align=ALIGN_CENTER, merge_ke=5)

    info_proyek = [
        ("Nama Proyek",     nama_proyek),
        ("No. Laporan",     nomor_laporan),
        ("Tanggal Hitung",  datetime.now().strftime("%d %B %Y")),
        ("Dibuat Oleh",     nama_konsultan or "—"),
        ("Acuan Standar",   "SNI 8460:2017 · Meyerhof (1976) · Tomlinson (2008)"),
    ]
    for i, (label, nilai) in enumerate(info_proyek):
        _tambah_info_baris(ws, 3 + i, label, nilai)

    # Spasi
    ws.append([])

    # Blok parameter tiang
    baris_th = 9
    _style_sel(ws, baris_th, 1, nilai="PARAMETER TIANG",
               font=FONT_HEADER, fill=FILL_HEADER,
               align=ALIGN_CENTER, merge_ke=5)

    tiang_data = [
        ("Tipe tiang",              param_tiang["tipe"]),
        ("Dimensi",                 param_tiang["dim_label"]),
        ("Kedalaman tiang (L)",     f"{param_tiang['kedalaman']:.2f} m"),
        ("Luas ujung (Ab)",         f"{param_tiang['area_ujung']:.4f} m²"),
        ("Keliling tiang (p)",      f"{param_tiang['keliling']:.4f} m"),
        ("Muka air tanah (MAT)",    f"{param_tiang['muka_air']:.2f} m"),
        ("Material",                param_tiang["material"]),
        ("Safety factor tekan",     f"{param_tiang['sf_tekan']:.1f}"),
        ("Safety factor tarik",     f"{param_tiang['sf_tarik']:.1f}"),
        ("Tipe displacement",       "Ya (displacement)" if param_tiang["is_displacement"] else "Tidak (non-displacement)"),
    ]
    for i, (label, nilai) in enumerate(tiang_data):
        _tambah_info_baris(ws, baris_th + 1 + i, label, nilai)

    # Catatan komersial
    baris_cat = baris_th + len(tiang_data) + 3
    _style_sel(ws, baris_cat, 1,
               nilai="CATATAN: File ini berisi nilai hasil perhitungan. "
                     "Rumus perhitungan tersimpan di sistem dan tidak ditampilkan di sini.",
               font=FONT_KECIL, align=ALIGN_LEFT, merge_ke=5)


# ==============================================================
# SHEET 2: PROFIL TANAH
# ==============================================================

def _buat_sheet_tanah(ws, df_tanah, param_tiang: dict) -> None:
    from utils.input_handler import KOLOM_TANAH
    ws.title = "2_Profil_Tanah"
    _set_lebar_kolom(ws, {
        "A": 6, "B": 22, "C": 14, "D": 14,
        "E": 10, "F": 10, "G": 12, "H": 10, "I": 12
    })

    _tambah_judul_sheet(ws, "PROFIL TANAH & DATA SPT", 1, 1, 9)

    _baris_header(ws, 2, 1, [
        "No.", "Jenis Tanah", "z Atas\n(m)", "z Bawah\n(m)",
        "Tebal\n(m)", "SPT-N", "Cu\n(kPa)", "φ\n(°)", "γ\n(kN/m³)"
    ])
    ws.row_dimensions[2].height = 30

    for i, (_, baris) in enumerate(df_tanah.iterrows()):
        tebal = float(baris[KOLOM_TANAH["z_bawah"]]) - float(baris[KOLOM_TANAH["z_atas"]])
        _baris_data(ws, 3 + i, 1, [
            i + 1,
            str(baris[KOLOM_TANAH["jenis"]]),
            float(baris[KOLOM_TANAH["z_atas"]]),
            float(baris[KOLOM_TANAH["z_bawah"]]),
            round(tebal, 2),
            int(baris[KOLOM_TANAH["spt"]]),
            float(baris[KOLOM_TANAH["cu"]]),
            float(baris[KOLOM_TANAH["phi"]]),
            float(baris[KOLOM_TANAH["gamma"]]),
        ], rata_kanan=[0, 2, 3, 4, 5, 6, 7, 8], nomor_baris=i)

    # Garis muka air tanah sebagai catatan
    baris_cat = 3 + len(df_tanah) + 1
    _style_sel(ws, baris_cat, 1,
               nilai=f"Kedalaman muka air tanah (MAT) = {param_tiang['muka_air']:.2f} m",
               font=FONT_SUBHDR, align=ALIGN_LEFT, merge_ke=9)


# ==============================================================
# SHEET 3: HASIL TEKAN & TARIK
# ==============================================================

def _buat_sheet_tekan_tarik(ws, hasil_tekan: dict, param_tiang: dict) -> None:
    ws.title = "3_Hasil_Tekan_Tarik"
    _set_lebar_kolom(ws, {
        "A": 6, "B": 22, "C": 11, "D": 11, "E": 10,
        "F": 20, "G": 12, "H": 10, "I": 11, "J": 12
    })

    # --- Sub-judul: Distribusi skin friction ---
    _tambah_judul_sheet(ws, "DISTRIBUSI DAYA DUKUNG SELIMUT (SKIN FRICTION) PER LAPISAN", 1, 1, 10)

    _baris_header(ws, 2, 1, [
        "No.", "Jenis Tanah", "z Atas\n(m)", "z Bawah\n(m)", "Tebal\n(m)",
        "Metode", "σ'v\n(kPa)", "α / β", "fs\n(kPa)", "Qs\n(kN)"
    ])
    ws.row_dimensions[2].height = 30

    for i, lap in enumerate(hasil_tekan["detail_lapisan"]):
        alfa_beta = (round(lap["alpha"], 4) if lap["kategori"] == "lempung"
                     else round(lap["beta"], 4))
        _baris_data(ws, 3 + i, 1, [
            lap["no"],
            lap["jenis"],
            round(lap["z_atas"], 2),
            round(lap["z_bawah"], 2),
            round(lap["tebal"], 2),
            lap["metode"].split("(")[0].strip(),
            round(lap["sigma_v_eff"], 2),
            alfa_beta,
            round(lap["fs"], 3),
            round(lap["Qs"], 2),
        ], rata_kanan=[0, 2, 3, 4, 6, 7, 8, 9], nomor_baris=i)

    # Baris total skin friction
    baris_tot = 3 + len(hasil_tekan["detail_lapisan"])
    _style_sel(ws, baris_tot, 1, nilai="TOTAL",
               font=Font(name="Arial", bold=True, size=10),
               fill=FILL_KUNING, align=ALIGN_CENTER,
               border=BORDER_ALL, merge_ke=9)
    _style_sel(ws, baris_tot, 10,
               nilai=round(hasil_tekan["Qskin"], 2),
               font=Font(name="Arial", bold=True, size=10, color=HIJAU_TUA),
               fill=FILL_KUNING, align=ALIGN_RIGHT,
               border=BORDER_ALL, fmt=FMT_ANGKA2)

    # --- Sub-judul: Ringkasan kapasitas ---
    baris_rng = baris_tot + 2
    _style_sel(ws, baris_rng, 1, nilai="RINGKASAN DAYA DUKUNG",
               font=FONT_HEADER, fill=FILL_HEADER,
               align=ALIGN_CENTER, merge_ke=10)

    _baris_header(ws, baris_rng + 1, 1, [
        "Komponen", "Qultimit\n(kN)", "SF", "Qijin\n(kN)",
        "Keterangan", "", "", "", "", ""
    ])

    ringkasan = [
        ("End Bearing (Qpoint)",
         hasil_tekan["Qpoint"], "—", "—", "Meyerhof (1976)"),
        ("Skin Friction (ΣQskin)",
         hasil_tekan["Qskin"], "—", "—", "α-method / β-method"),
        ("Daya Dukung TEKAN",
         hasil_tekan["Qult_tekan"],
         hasil_tekan["sf_tekan"],
         hasil_tekan["Qijin_tekan"],
         "Qijin = Qult / SF"),
        ("Daya Dukung TARIK",
         hasil_tekan["Qult_tarik"],
         hasil_tekan["sf_tarik"],
         hasil_tekan["Qijin_tarik"],
         "Qijin_tarik = ΣQskin × fr / SF"),
    ]
    if hasil_tekan.get("Pn_struktur"):
        ringkasan.append((
            "Kapasitas Struktur (Pn)",
            hasil_tekan["Pn_struktur"], "—", "—",
            "SNI 2847:2019"
        ))

    for i, (label, qult, sf, qijin, ket) in enumerate(ringkasan):
        br = baris_rng + 2 + i
        fill = FILL_KUNING if "TEKAN" in label or "TARIK" in label else (
               FILL_GENAP if i % 2 == 0 else FILL_GANJIL)
        _style_sel(ws, br, 1, nilai=label,
                   font=Font(name="Arial", bold=("TEKAN" in label or "TARIK" in label), size=10),
                   fill=fill, align=ALIGN_LEFT, border=BORDER_ALL)
        _style_sel(ws, br, 2, nilai=qult if isinstance(qult, (int, float)) else qult,
                   font=FONT_ANGKA, fill=fill, align=ALIGN_RIGHT,
                   border=BORDER_ALL, fmt=FMT_ANGKA2)
        _style_sel(ws, br, 3, nilai=sf,
                   font=FONT_ANGKA, fill=fill, align=ALIGN_CENTER, border=BORDER_ALL)
        _style_sel(ws, br, 4,
                   nilai=qijin if isinstance(qijin, (int, float)) else qijin,
                   font=Font(name="Arial", bold=True, size=10, color=HIJAU_TUA)
                        if isinstance(qijin, (int, float)) else FONT_BODY,
                   fill=fill, align=ALIGN_RIGHT,
                   border=BORDER_ALL, fmt=FMT_ANGKA2 if isinstance(qijin, float) else None)
        _style_sel(ws, br, 5, nilai=ket,
                   font=FONT_KECIL, fill=fill, align=ALIGN_LEFT,
                   border=BORDER_ALL, merge_ke=10)


# ==============================================================
# SHEET 4: GAYA LATERAL
# ==============================================================

def _buat_sheet_lateral(ws, hasil_lateral: dict | None,
                         metode_lateral: str) -> None:
    ws.title = "4_Gaya_Lateral"
    _set_lebar_kolom(ws, {"A": 30, "B": 20, "C": 15, "D": 20})

    if not hasil_lateral:
        _tambah_judul_sheet(ws, "GAYA LATERAL — Belum dihitung", 1, 1, 4)
        _style_sel(ws, 3, 1,
                   nilai="Hitung gaya lateral di tab '↔️ Gaya Lateral' terlebih dahulu.",
                   font=FONT_KECIL, align=ALIGN_LEFT, merge_ke=4)
        return

    _tambah_judul_sheet(ws, f"GAYA LATERAL — Metode: {metode_lateral}", 1, 1, 4)

    _baris_header(ws, 2, 1, ["Parameter", "Nilai", "Satuan", "Keterangan"])

    if "Broms" in metode_lateral or "Hu" in hasil_lateral:
        data_lat = [
            ("Kapasitas lateral ultimit (Hu)", hasil_lateral.get("Hu", 0), "kN", "Broms (1964)"),
            ("Kapasitas lateral ijin (Hijin)",  hasil_lateral.get("Hijin", 0), "kN", "SF = 2.5"),
            ("Momen maksimum (Mmax)",           hasil_lateral.get("Mmax", 0), "kN·m", ""),
            ("Defleksi kepala tiang",           hasil_lateral.get("defleksi_mm", 0), "mm", "Elastis"),
            ("Kontrol (H ≤ Hijin)",
             "OK" if hasil_lateral.get("kontrol_ok") else "TIDAK OK", "—", ""),
            ("Tanah dominan",                   hasil_lateral.get("tanah_dominan", "—"), "—", ""),
            ("Kondisi kepala",                  hasil_lateral.get("kondisi", "—"), "—", ""),
        ]
    else:  # P-Y curve
        data_lat = [
            ("Defleksi kepala tiang (y₀)", hasil_lateral.get("y0_mm", 0), "mm", "P-Y curve FD"),
            ("Momen maksimum (Mmax)",       hasil_lateral.get("Mmax", 0), "kN·m", ""),
            ("z titik Mmax",               hasil_lateral.get("z_Mmax", 0), "m", ""),
            ("Gaya geser maks (Vmax)",      hasil_lateral.get("Vmax", 0), "kN", ""),
            ("Iterasi konvergensi",
             f"{'Konvergen' if hasil_lateral.get('konvergen') else 'Belum konvergen'} ({hasil_lateral.get('iterasi',0)} iter)",
             "—", ""),
            ("EI tiang",                   hasil_lateral.get("EI", 0), "kN·m²", ""),
        ]

    for i, baris in enumerate(data_lat):
        fill = FILL_GENAP if i % 2 == 0 else FILL_GANJIL
        _style_sel(ws, 3 + i, 1, nilai=baris[0],
                   font=FONT_SUBHDR, fill=FILL_SUBHDR,
                   align=ALIGN_LEFT, border=BORDER_ALL)
        # Nilai: angka atau teks
        val = baris[1]
        fmt_val = FMT_ANGKA2 if isinstance(val, float) else None
        _style_sel(ws, 3 + i, 2, nilai=val,
                   font=FONT_ANGKA, fill=fill, align=ALIGN_RIGHT,
                   border=BORDER_ALL, fmt=fmt_val)
        _style_sel(ws, 3 + i, 3, nilai=baris[2],
                   font=FONT_BODY, fill=fill, align=ALIGN_CENTER, border=BORDER_ALL)
        _style_sel(ws, 3 + i, 4, nilai=baris[3],
                   font=FONT_KECIL, fill=fill, align=ALIGN_LEFT, border=BORDER_ALL)

    # Jika p-y curve, tambahkan tabel profil defleksi (setiap 5 node)
    if "y_m" in hasil_lateral:
        baris_tbl = 3 + len(data_lat) + 2
        _style_sel(ws, baris_tbl, 1, nilai="PROFIL DEFLEKSI, MOMEN & GESER",
                   font=FONT_HEADER, fill=FILL_HEADER,
                   align=ALIGN_CENTER, merge_ke=4)
        _baris_header(ws, baris_tbl + 1, 1, ["z (m)", "y (mm)", "M (kN·m)", "V (kN)"])
        z_list = hasil_lateral["z_nodes"]
        y_list = hasil_lateral["y_m"]
        M_list = hasil_lateral["M_kNm"]
        V_list = hasil_lateral["V_kN"]
        step   = max(1, len(z_list) // 20)  # maks ~20 baris tabel
        for idx, i in enumerate(range(0, len(z_list), step)):
            _baris_data(ws, baris_tbl + 2 + idx, 1, [
                round(z_list[i], 2),
                round(y_list[i] * 1000, 3),
                round(M_list[i], 2),
                round(V_list[i], 2),
            ], rata_kanan=[0, 1, 2, 3], nomor_baris=idx)


# ==============================================================
# SHEET 5: GRAFIK
# ==============================================================

def _buat_sheet_grafik(ws, param_tiang: dict, df_tanah,
                        hasil_tekan: dict,
                        hasil_lateral: dict | None = None,
                        metode_lateral: str = "") -> None:
    from utils.grapher import (
        buat_grafik_profil,
        buat_grafik_distribusi_skin,
        buat_grafik_variasi_kedalaman,
    )
    from calculations.bearing_capacity import hitung_variasi_kedalaman

    ws.title = "5_Grafik"
    _tambah_judul_sheet(ws, "GRAFIK HASIL PERHITUNGAN", 1, 1, 10)
    _style_sel(ws, 2, 1,
               nilai="Grafik diekspor dari hasil perhitungan. "
                     "Tidak dapat diedit langsung di Excel.",
               font=FONT_KECIL, align=ALIGN_LEFT, merge_ke=10)

    baris_gambar = 4

    def _sisipkan_grafik(fig, baris_mulai, lebar_px=480, tinggi_px=360):
        """Menyisipkan grafik matplotlib ke worksheet."""
        buf = _gambar_ke_bytes(fig, dpi=120)
        img = XLImage(buf)
        img.width  = lebar_px
        img.height = tinggi_px
        anchor_sel = f"A{baris_mulai}"
        ws.add_image(img, anchor_sel)
        # Perkiraan baris yang terpakai (~tinggi / 15px per baris)
        return baris_mulai + int(tinggi_px / 15) + 2

    # Grafik 1: Profil tanah
    _style_sel(ws, baris_gambar - 1, 1, nilai="Grafik 1: Profil Tanah & SPT-N",
               font=FONT_SUBHDR, align=ALIGN_LEFT)
    fig1 = buat_grafik_profil(
        hasil_tekan["semua_lapisan"],
        param_tiang["kedalaman"],
        param_tiang["muka_air"],
        param_tiang["diameter"]
    )
    baris_gambar = _sisipkan_grafik(fig1, baris_gambar, 360, 420)

    # Grafik 2: Distribusi skin friction
    _style_sel(ws, baris_gambar - 1, 1,
               nilai="Grafik 2: Distribusi Skin Friction per Lapisan",
               font=FONT_SUBHDR, align=ALIGN_LEFT)
    fig2 = buat_grafik_distribusi_skin(
        hasil_tekan["detail_lapisan"],
        hasil_tekan["Qpoint"],
        param_tiang["kedalaman"]
    )
    baris_gambar = _sisipkan_grafik(fig2, baris_gambar, 540, 360)

    # Grafik 3: Variasi kedalaman
    hasil_var = hitung_variasi_kedalaman(
        df_tanah, param_tiang,
        z_min=max(param_tiang["kedalaman"] * 0.3, 3.0),
        z_max=param_tiang["kedalaman"]
    )
    if hasil_var:
        _style_sel(ws, baris_gambar - 1, 1,
                   nilai="Grafik 3: Kapasitas Tiang vs Variasi Kedalaman",
                   font=FONT_SUBHDR, align=ALIGN_LEFT)
        fig3 = buat_grafik_variasi_kedalaman(hasil_var)
        baris_gambar = _sisipkan_grafik(fig3, baris_gambar, 480, 360)

    # Grafik 4: Gaya lateral
    if hasil_lateral:
        from utils.grapher_lateral import buat_grafik_broms, buat_grafik_py
        _style_sel(ws, baris_gambar - 1, 1,
                   nilai=f"Grafik 4: Gaya Lateral ({metode_lateral})",
                   font=FONT_SUBHDR, align=ALIGN_LEFT)
        if "Broms" in metode_lateral or "Hu" in hasil_lateral:
            hasil_lateral["H_input"] = hasil_lateral.get("H_input", 0)
            hasil_lateral["L"] = param_tiang["kedalaman"]
            hasil_lateral["D"] = param_tiang["diameter"]
            fig4 = buat_grafik_broms(hasil_lateral)
            _sisipkan_grafik(fig4, baris_gambar, 540, 360)
        elif "y_m" in hasil_lateral:
            fig4 = buat_grafik_py(hasil_lateral, param_tiang, df_tanah)
            _sisipkan_grafik(fig4, baris_gambar, 720, 480)


# ==============================================================
# FUNGSI UTAMA: BUAT FILE EXCEL
# ==============================================================

def buat_excel(
    param_tiang: dict,
    df_tanah,
    hasil_tekan: dict,
    hasil_lateral: dict | None = None,
    metode_lateral: str = "—",
    nama_proyek: str = "Proyek Pondasi",
    nama_konsultan: str = "",
    nomor_laporan: str = "LAP-001",
    sertakan_grafik: bool = True,
) -> io.BytesIO:
    """
    Membuat file Excel lengkap berisi 4–5 sheet.
    HANYA nilai, TIDAK ADA formula — untuk tujuan komersial.

    Mengembalikan BytesIO siap di-download.
    """
    wb = Workbook()
    # Hapus sheet default
    wb.remove(wb.active)

    # Buat 5 sheet
    ws1 = wb.create_sheet()
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet() if sertakan_grafik else None

    _buat_sheet_info(ws1, param_tiang, nama_proyek, nama_konsultan, nomor_laporan)
    _buat_sheet_tanah(ws2, df_tanah, param_tiang)
    _buat_sheet_tekan_tarik(ws3, hasil_tekan, param_tiang)
    _buat_sheet_lateral(ws4, hasil_lateral, metode_lateral)
    if ws5 and sertakan_grafik:
        _buat_sheet_grafik(ws5, param_tiang, df_tanah,
                           hasil_tekan, hasil_lateral, metode_lateral)

    # Aktifkan sheet pertama saat dibuka
    wb.active = ws1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
